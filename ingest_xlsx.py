from __future__ import annotations

import argparse
import dataclasses
import datetime as dt
from collections import Counter
import os
import re
import sqlite3
from typing import Any, Iterable

import openpyxl
import pandas as pd
import yaml
from rapidfuzz import fuzz, process

from recipient_identity import assign_recipient_ids
from openpyxl.utils.datetime import from_excel as _excel_serial_to_datetime


EXPECTED_COLS = 13  # current vendor sheet has 13 columns (incl. address/phone/request)

FILENAME_DATE_RE = re.compile(r"^(?P<yymmdd>\d{6})")
SIZE_RE = re.compile(r"\bW\s*(\d{2,5})\s*[*xX]\s*D\s*(\d{2,5})(?:\s*[*xX]\s*H\s*(\d{2,5}))?\b")
CM_RE = re.compile(r"\b(\d{2,3})\s*cm\b", re.IGNORECASE)

COLOR_WORDS = [
    "크림화이트",
    "크림",
    "오프화이트",
    "화이트",
    "블랙",
    "그레이",
    "월넛",
    "오크",
    "베이지",
    "브라운",
    "내추럴",
    "자작",
    # 원단/시트 상품명 (규격 열은 cm만 있고 색 열에만 적히는 경우)
    "올드블루진스",
    "런던브릭",
    # 원단/컬러 추가 (실제 주문서에서 다리/옵션 칸에만 들어오는 케이스 보정)
    "카멜리아",
    "소프트세이지",
    "옥스포드옐로우",
    "세이지",
    "무슬린",
    "소프트퍼플",
    "윈터선샤인",
    "다크네이비",
    "클래식블루",
    "로즈로코코",
]


@dataclasses.dataclass(frozen=True)
class ItemRow:
    source_file: str
    row_idx: int
    deadline_raw: str | None
    group_no: str | None
    group_start_row: int | None
    receiver_name_raw: str | None
    order_date_raw: str | None
    product_raw: str | None
    spec_raw: str | None
    shelf_color_raw: str | None
    leg_color_raw: str | None
    qty_raw: str | None
    ship_raw: str | None
    address_raw: str | None
    phone_raw: str | None
    delivery_request_raw: str | None
    attention_note_raw: str | None


def purchase_date_from_filename(source_file: str) -> str | None:
    """
    Extract purchase date from file prefix like: 260401-02_....xlsx -> 2026-04-01
    """
    m = FILENAME_DATE_RE.match(source_file)
    if not m:
        return None
    yymmdd = m.group("yymmdd")
    try:
        yy = int(yymmdd[0:2])
        mm = int(yymmdd[2:4])
        dd = int(yymmdd[4:6])
        yyyy = 2000 + yy
        return dt.date(yyyy, mm, dd).isoformat()
    except Exception:
        return None


def parse_purchase_date_from_sheet(raw: object, source_file: str | None = None) -> str | None:
    """
    엑셀의 '주문일자' 셀 값을 ISO 날짜(YYYY-MM-DD)로 변환합니다.
    파일명(YYMMDD)보다 우선해서 접수일자를 날짜별로 정확히 구분하기 위함입니다.
    """
    if raw is None:
        return None
    if isinstance(raw, dt.datetime):
        try:
            return raw.date().isoformat()
        except Exception:
            return None
    if isinstance(raw, dt.date):
        return raw.isoformat()

    s = str(raw).strip()
    if not s or s.lower() in ("none", "nan", "-", "#n/a"):
        return None

    # "2026-04-16 00:00:00" / "2026-04-16"
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        try:
            return dt.date.fromisoformat(s[:10]).isoformat()
        except ValueError:
            pass

    # "4.19" / "4/19" 같은 형태 (대부분 주문서에서 사용하는 월.일)
    m = re.fullmatch(r"(?P<m>\d{1,2})\s*[./-]\s*(?P<d>\d{1,2})", s)
    if m:
        mm = int(m.group("m"))
        dd = int(m.group("d"))
        # 연도는 파일명 YYMMDD에서 추정. (없으면 올해)
        base_iso = purchase_date_from_filename(source_file or "") if source_file else None
        base = None
        try:
            base = dt.date.fromisoformat(base_iso) if base_iso else None
        except Exception:
            base = None
        yyyy = (base.year if base else dt.date.today().year)
        # 12월 주문이 4월 파일에 섞이는 케이스 등: base 월보다 크게 멀면 작년으로 간주
        if base and (mm - base.month) >= 6:
            yyyy = base.year - 1
        try:
            return dt.date(yyyy, mm, dd).isoformat()
        except Exception:
            return None

    ts = pd.to_datetime(s, errors="coerce", dayfirst=False)
    if not pd.isna(ts):
        try:
            return ts.date().isoformat()
        except Exception:
            pass

    # 숫자만(엑셀 날짜 직렬값) — openpyxl data_only 시 datetime이 아닌 경우
    if re.fullmatch(r"\d{5,6}", s):
        try:
            n = float(s)
            if 30000 <= n <= 80000:
                d = _excel_serial_to_datetime(n)
                if isinstance(d, dt.datetime):
                    return d.date().isoformat()
                if isinstance(d, dt.date):
                    return d.isoformat()
        except Exception:
            pass

    return None


def extract_size(text: str | None) -> str | None:
    if not text:
        return None
    t = text.replace("\n", " ")
    m = SIZE_RE.search(t)
    if m:
        w, d, h = m.group(1), m.group(2), m.group(3)
        return f"W{w}*D{d}" + (f"*H{h}" if h else "")
    m2 = CM_RE.search(t)
    if m2:
        return f"{m2.group(1)}cm"
    return None


def _first_color(text: str | None) -> str | None:
    if not text:
        return None
    t = text.replace("\n", " ")
    for c in COLOR_WORDS:
        if c in t:
            return c
    return None


def _shelf_color_fallback_from_leg_cell(leg_cell: str | None) -> str | None:
    """책장색 열이 비었을 때, '다리' 열에 실제로는 원단/컬러만 적힌 행(예: '올드블루진스 / 25cm') 보정."""
    if not leg_cell:
        return None
    t = str(leg_cell).replace("\n", " ")
    if re.search(r"다리", t):
        return None
    return _first_color(t)


def shelf_color_from_note_raw_cell(raw: object) -> str | None:
    """DB `note_raw`(엑셀 다리색 열 원본)만 보고 책장색 보정 — 인제스트·대시보드가 동일 규칙을 쓰게 한다."""
    if raw is None:
        return None
    try:
        if pd.isna(raw):
            return None
    except Exception:
        pass
    s = str(raw).strip()
    if not s or s.lower() in ("none", "nan", "-", "#n/a"):
        return None
    return _shelf_color_fallback_from_leg_cell(s)


def extract_leg_color(text: str | None) -> str | None:
    if not text:
        return None
    t = text.replace("\n", " ")
    # try "다리: 블랙" / "다리 블랙" patterns first
    m = re.search(r"다리\s*[:\-]?\s*([가-힣]+)", t)
    if m:
        cand = m.group(1)
        for c in COLOR_WORDS:
            if c in cand:
                return c
    return None


def _to_str(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return str(v).strip() or None


def _looks_like_int(s: str | None) -> bool:
    if not s:
        return False
    return bool(re.fullmatch(r"\d+", s.strip()))


def _parse_excel(path: str) -> list[ItemRow]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Heuristic: header is first row with many non-empty cells.
    header_row = None
    for r in range(1, 21):
        vals = [_to_str(ws.cell(r, c).value) for c in range(1, EXPECTED_COLS + 1)]
        nonempty = sum(1 for v in vals if v)
        if nonempty >= 7:
            header_row = r
            break
    if header_row is None:
        header_row = 1

    rows: list[ItemRow] = []
    current_deadline = None
    current_group_no = None
    current_group_start_row: int | None = None
    current_receiver_name = None
    current_order_date = None
    current_address = None
    current_phone = None
    current_delivery_request = None
    current_attention_note = None

    # Some order sheets contain an "instruction/note" in column A (e.g. A2/A9)
    # like: "(3/28 주문 서혜선님 제품변경, 컬러지정하였습니다. ...)"
    # We capture these and attach them to the matching receiver as an attention note.
    pending_note_by_name: dict[str, str] = {}
    pending_unassigned_notes: list[str] = []

    def _extract_name_note(s: str | None) -> tuple[str | None, str | None]:
        if not s:
            return None, None
        s2 = str(s).strip()
        if not s2:
            return None, None
        # Typical keywords that appear in these attention notes.
        # Helps avoid false positives when scanning a lot of rows.
        if not re.search(r"(변경|컬러|색상|추가|수정|주소|옵션|취소|반품|교환)", s2):
            return None, None

        # Many sheets start these lines with a date like "3/28 ..."
        # Use it as an extra safety signal when order keywords are absent.
        has_order_kw = bool(re.search(r"(주문|주문자)", s2))
        has_date_prefix = bool(re.match(r"^\s*\d{1,2}\s*/\s*\d{1,2}", s2))
        if not has_order_kw and not has_date_prefix:
            # Allow lines like: "장재국 색상지정 ..." (no date/order keyword)
            m0 = re.search(r"([가-힣]{2,10}).*(?:색상|컬러|색깔).*(?:지정)", s2)
            if m0:
                nm0 = (m0.group(1) or "").strip()
                return (nm0 or None), s2
            return None, None

        # Primary: "주문 홍길동님" or "주문 홍길동 ..."
        m = re.search(r"(?:주문|주문자)\s*([가-힣]{2,10})(?:\s*님)?", s2)
        if m:
            name = (m.group(1) or "").strip()
            if name:
                return name, s2

        # Fallback: if it contains "님", grab the nearest Hangul name before it.
        m2 = re.search(r"([가-힣]{2,10})\s*님", s2)
        if m2:
            name = (m2.group(1) or "").strip()
            if name:
                return name, s2

        # Keep as unassigned note (we may still match by inclusion later)
        return None, s2

    def _merge_req(existing: str | None, extra: str | None) -> str | None:
        ex = (existing or "").strip()
        ex2 = (extra or "").strip()
        if not ex and not ex2:
            return None
        if not ex:
            return ex2
        if not ex2:
            return ex
        if ex2 in ex:
            return ex
        # Keep as a single line for UI scanning
        return f"{ex} / {ex2}"

    # Scan A-column notes across the sheet (some files put them far below, e.g. A106).
    # Keep it bounded to avoid pathological sheets.
    scan_end = min(ws.max_row, header_row + 500)
    for r in range(header_row + 1, scan_end + 1):
        note = _to_str(ws.cell(r, 1).value)
        nm, txt = _extract_name_note(note)
        if nm and txt:
            pending_note_by_name[nm] = _merge_req(pending_note_by_name.get(nm), txt) or txt
        elif txt:
            # Keep as a fallback: later we can match it by inclusion against receiver_name.
            pending_unassigned_notes.append(txt)

    def _pick_note_for_receiver(receiver: str | None) -> str | None:
        if not receiver:
            return None
        rn = str(receiver).strip()
        if not rn:
            return None
        # Exact match first
        if rn in pending_note_by_name:
            return pending_note_by_name[rn]
        # Fallback: containment match (handles cases like "서혜선(네이버)" vs "서혜선")
        for k, v in pending_note_by_name.items():
            kk = (k or "").strip()
            if not kk:
                continue
            if kk in rn or rn in kk:
                return v
        # Final fallback: match unassigned note text that contains the receiver name.
        rn2 = re.sub(r"\s+", "", rn)
        for txt in pending_unassigned_notes:
            t2 = re.sub(r"\s+", "", txt or "")
            if rn2 and t2 and (rn2 in t2):
                return txt
        return None

    blank_run = 0
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [_to_str(ws.cell(r, c).value) for c in range(1, EXPECTED_COLS + 1)]
        if all(v is None for v in vals):
            blank_run += 1
            if blank_run >= 10:
                break
            continue
        blank_run = 0

        (
            deadline,
            group_no,
            receiver_name,
            order_date,
            product,
            spec,
            shelf_color,
            leg_color,
            qty,
            ship,
            address,
            phone,
            delivery_request,
        ) = vals

        # Rows where "번호" is a digit: may be a new order block, or the same order with the
        # number repeated on every line. Only advance group_start_row when the block actually
        # changes (new 번호, or same 번호 but a new 받는분), so one customer does not split into
        # multiple order_id values.
        if _looks_like_int(group_no):

            def _norm_group_digits(s: str | None) -> str | None:
                if not s or not _looks_like_int(s):
                    return None
                return str(int(str(s).strip()))

            new_gn = _norm_group_digits(group_no)
            old_gn = _norm_group_digits(str(current_group_no)) if current_group_no is not None else None
            recv_cell = _to_str(receiver_name)
            prev_recv = _to_str(current_receiver_name)

            advance_start = old_gn is None or new_gn != old_gn
            if not advance_start and recv_cell and prev_recv and recv_cell.strip() != prev_recv.strip():
                advance_start = True
            if not advance_start and recv_cell and not prev_recv:
                advance_start = True

            if advance_start:
                current_group_start_row = r

            current_deadline = deadline or current_deadline
            current_group_no = group_no
            current_receiver_name = receiver_name or current_receiver_name
            current_order_date = order_date or current_order_date
            current_address = address or current_address
            current_phone = phone or current_phone
            # delivery_request is order-specific; DO NOT carry over from previous groups.
            if advance_start:
                current_delivery_request = delivery_request
            else:
                dr = _to_str(delivery_request)
                if dr:
                    current_delivery_request = delivery_request

            # Attach any pending A-column note that matches this receiver name.
            current_attention_note = None
            extra = _pick_note_for_receiver(current_receiver_name)
            if extra:
                current_attention_note = extra

        # Item rows: product exists OR spec/option exists AND we have an active group.
        is_itemish = bool(product or spec or shelf_color or leg_color) and bool(current_group_no)
        if not is_itemish:
            continue

        rows.append(
            ItemRow(
                source_file=os.path.basename(path),
                row_idx=r,
                deadline_raw=deadline or current_deadline,
                group_no=current_group_no,
                group_start_row=current_group_start_row,
                receiver_name_raw=receiver_name or current_receiver_name,
                order_date_raw=order_date or current_order_date,
                product_raw=product,
                spec_raw=spec,
                shelf_color_raw=shelf_color,
                leg_color_raw=leg_color,
                qty_raw=qty,
                ship_raw=ship,
                address_raw=address or current_address,
                phone_raw=phone or current_phone,
                delivery_request_raw=delivery_request or current_delivery_request,
                attention_note_raw=current_attention_note,
            )
        )

    return rows


def _normalize_text(s: str) -> str:
    s = s.strip().lower()
    s = s.replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s)
    return s


def _product_key(s: str) -> str:
    s = _normalize_text(s)
    # remove common option markers but keep meaningful words/numbers
    s = s.replace("\\n", " ")
    s = re.sub(r"\([^)]*\)", " ", s)  # remove (...) blocks like (L)
    s = re.sub(r"\[[^]]*\]", " ", s)  # remove [...] blocks
    s = re.sub(r"\s+", " ", s).strip()
    return s


def load_alias_map(path: str) -> tuple[dict[str, str], list[str]]:
    """
    Returns:
      - alias_key -> canonical
      - list of canonicals
    """
    if not os.path.exists(path):
        return {}, []
    with open(path, "r", encoding="utf-8") as f:
        raw = yaml.safe_load(f) or {}
    alias_to_canon: dict[str, str] = {}
    canonicals: list[str] = []
    for canon, aliases in raw.items():
        canon = (canon or "").strip()
        if not canon:
            continue
        canonicals.append(canon)
        alias_to_canon[_product_key(canon)] = canon
        for a in (aliases or []):
            a = (a or "").strip()
            if not a:
                continue
            alias_to_canon[_product_key(a)] = canon
    return alias_to_canon, sorted(set(canonicals))


def resolve_product(
    product_raw: str | None,
    alias_to_canon: dict[str, str],
    canonicals: list[str],
) -> tuple[str | None, str | None, float | None]:
    if not product_raw:
        return None, None, None
    k = _product_key(product_raw)
    if not k:
        return None, None, None
    if k in alias_to_canon:
        return alias_to_canon[k], None, 100.0

    # Fuzzy suggestion against canonical list (by normalized key)
    if not canonicals:
        return None, None, None
    canon_keys = {_product_key(c): c for c in canonicals}
    hit = process.extractOne(
        k,
        list(canon_keys.keys()),
        scorer=fuzz.WRatio,
    )
    if not hit:
        return None, None, None
    best_key, score, _idx = hit
    return None, canon_keys.get(best_key), float(score)


def _to_int(s: str | None) -> int | None:
    if not s:
        return None
    s2 = re.sub(r"[^\d]", "", s)
    return int(s2) if s2 else None


# 엑셀 품목 `ship_raw` → 직접/택배 판별(주문 다수결에 사용).
SETTLEMENT_SHIP_VALUES: tuple[str, ...] = ("직접배송", "택배")


def classify_ship_raw(raw: object) -> str | None:
    """한 줄의 ship_raw를 직접배송/택배로 분류. 알 수 없으면 None."""
    if raw is None:
        return None
    try:
        if pd.isna(raw):
            return None
    except Exception:
        pass
    s = _normalize_text(str(raw))
    if not s:
        return None
    # 엑셀에 '직접' / '택배'만 적는 경우(가장 흔함)
    compact = s.replace(" ", "")
    if compact in ("직접", "직접."):
        return "직접배송"
    if compact in ("택배", "택배."):
        return "택배"
    direct_markers = (
        "직접배송",
        "직접 배송",
        "직접배",
        "직배",
        "직송",
        "직배송",
        "직배차",
        "방문수령",
        "방문 수령",
        "매장수령",
        "매장 수령",
        "직거래",
        "자체배송",
        "자체 배송",
        "자체픽업",
        "고객방문",
        "방문픽업",
        "방문 배송",
        "방문배송",
        "당사배송",
        "당사 배송",
        "지게차",
        "사다리차",
    )
    if any(k in s for k in direct_markers):
        return "직접배송"
    if any(k in s for k in ("택배", "로젠", "parcel", "courier")):
        return "택배"
    return None


def infer_settlement_ship_series(items_df: pd.DataFrame) -> pd.Series:
    """order_id -> '직접배송' | '택배' from 엑셀 품목 ship_raw (다수결; 동수·미기재는 택배)."""
    if items_df is None or len(items_df) == 0 or "order_id" not in items_df.columns:
        return pd.Series(dtype=str)
    tmp = items_df[["order_id", "ship_raw"]].copy()
    tmp["_kind"] = tmp["ship_raw"].map(classify_ship_raw)

    def _vote(series: pd.Series) -> str:
        vals = [v for v in series.dropna().tolist() if v in SETTLEMENT_SHIP_VALUES]
        if not vals:
            return "택배"
        c = Counter(vals)
        best_n = c.most_common(1)[0][1]
        tops = [k for k, v in c.items() if v == best_n]
        if len(tops) == 1:
            return tops[0]
        return "택배"

    out = tmp.groupby("order_id", sort=False)["_kind"].agg(_vote)
    # SQLite 등에서 order_id dtype이 섞이면 대시보드 map 시 전부 택배로 떨어질 수 있어 인덱스를 문자열로 고정
    out.index = out.index.astype(str)
    return out


def _as_text_blob(v: object) -> str:
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    s = str(v)
    if s.strip().lower() in ("nan", "none", "-", "#n/a"):
        return ""
    return s


def _sheet_text_blob_for_order(order_row: pd.Series, items_df: pd.DataFrame) -> str:
    """주문 1건 + 해당 품목 줄 전체에서 '도면참조' 검색용 텍스트."""
    oid = str(order_row.get("order_id") or "").strip()
    parts: list[str] = []
    for key in ("receiver_name", "delivery_request", "attention_note", "order_list", "special_issue", "deadline_raw"):
        if key in order_row.index:
            parts.append(_as_text_blob(order_row.get(key)))
    if oid and len(items_df) and "order_id" in items_df.columns:
        sub = items_df.loc[items_df["order_id"].astype(str) == oid]
        for _, ir in sub.iterrows():
            for key in ("product_raw", "spec_raw", "note_raw", "ship_raw", "product_canonical"):
                if key in ir.index:
                    parts.append(_as_text_blob(ir.get(key)))
    return "\n".join(parts)


def order_sheet_blob_for_drawing_ref(order_row: pd.Series, items_df: pd.DataFrame) -> str:
    """대시보드 이름 옆 📐 표시와 `apply_drawing_ref_status_from_sheet`와 동일한 본문 범위."""
    return _sheet_text_blob_for_order(order_row, items_df)


def sheet_contains_drawing_ref_keyword(text: str) -> bool:
    """엑셀·현장에서 '도면참조'와 '도면참고'(참고)를 같은 의도로 쓰는 경우가 많음."""
    if not text:
        return False
    compact = re.sub(r"[\s\u00a0]+", "", text)
    return "도면참조" in compact or "도면참고" in compact


def apply_drawing_ref_status_from_sheet(orders_df: pd.DataFrame, items_df: pd.DataFrame) -> pd.DataFrame:
    """엑셀에서 온 본문에 '도면참조'가 있으면 status를 도면참조로(접수에서만 승격)."""
    if orders_df is None or len(orders_df) == 0 or "status" not in orders_df.columns:
        return orders_df
    out = orders_df.copy()
    st = out["status"].astype(str).str.strip()
    hit: list[bool] = []
    for _, row in out.iterrows():
        hit.append(sheet_contains_drawing_ref_keyword(_sheet_text_blob_for_order(row, items_df)))
    m = pd.Series(hit, index=out.index).fillna(False) & st.eq("접수")
    out.loc[m, "status"] = "도면참조"
    return out


def build_frames(rows: Iterable[ItemRow], alias_path: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    alias_to_canon, canonicals = load_alias_map(alias_path)
    order_rows: dict[str, dict[str, Any]] = {}
    item_out: list[dict[str, Any]] = []
    unknown_products: dict[str, int] = {}

    for it in rows:
        # group_no can repeat inside the same sheet; include group_start_row to disambiguate.
        start = it.group_start_row if it.group_start_row is not None else it.row_idx
        order_id = f"{it.source_file}#{it.group_no}@{start}"
        if order_id not in order_rows:
            sheet_p = parse_purchase_date_from_sheet(it.order_date_raw, it.source_file)
            order_rows[order_id] = {
                "order_id": order_id,
                "source_file": it.source_file,
                "group_no": it.group_no,
                "group_start_row": it.group_start_row,
                "deadline_raw": it.deadline_raw,
                "order_date_raw": it.order_date_raw,
                # 구매일자: 엑셀 주문일자 우선, 없으면 파일명 YYMMDD (예: 260401-.. -> 2026-04-01)
                "purchase_date": sheet_p or purchase_date_from_filename(it.source_file),
                # 주문자(받는사람) / 배송 정보: 이 엑셀(3열, 11~13열)에서 추출
                "receiver_name": it.receiver_name_raw,
                "address": it.address_raw,
                "phone": it.phone_raw,
                "delivery_request": it.delivery_request_raw,
                # 엑셀 A열 특이사항(주문 ○○님 ...) 자동 반영
                "attention_note": it.attention_note_raw,
                # 주문목록(요약): items에서 생성
                "order_list": None,
                # 특이사항(리콜/클레임 등) - 수기 입력용
                "special_issue": None,
                "status": "접수",
                # 출고 처리 시각(대시보드에서 설정). 재수집 시 보존.
                "shipped_at": None,
                "created_at": dt.datetime.now().isoformat(timespec="seconds"),
            }
        else:
            # Fill missing order-level fields if later rows contain them
            o = order_rows[order_id]
            sheet_p = parse_purchase_date_from_sheet(it.order_date_raw, it.source_file)
            if sheet_p:
                o["purchase_date"] = sheet_p
            for k, v in [
                ("receiver_name", it.receiver_name_raw),
                ("address", it.address_raw),
                ("phone", it.phone_raw),
                ("delivery_request", it.delivery_request_raw),
                ("attention_note", it.attention_note_raw),
            ]:
                if (o.get(k) is None or str(o.get(k)).strip() == "") and v:
                    o[k] = v

        canon, suggestion, score = resolve_product(it.product_raw, alias_to_canon, canonicals)
        if canon is None and it.product_raw:
            unknown_products[_product_key(it.product_raw)] = unknown_products.get(_product_key(it.product_raw), 0) + 1

        item_out.append(
            {
                "order_id": order_id,
                "source_file": it.source_file,
                "row_idx": it.row_idx,
                "product_raw": it.product_raw,
                "product_key": _product_key(it.product_raw or "") if it.product_raw else None,
                "product_canonical": canon,
                "suggested_canonical": suggestion,
                "suggestion_score": score,
                "spec_raw": it.spec_raw,
                "note_raw": it.leg_color_raw,
                "size": extract_size(it.spec_raw) or extract_size(it.leg_color_raw) or extract_size(it.product_raw),
                # Prefer explicit sheet columns when present; otherwise fallback to keyword extraction.
                "shelf_color": it.shelf_color_raw
                or _first_color(it.spec_raw)
                or _first_color(it.product_raw)
                or _shelf_color_fallback_from_leg_cell(it.leg_color_raw),
                "leg_color": extract_leg_color(it.leg_color_raw) or extract_leg_color(it.spec_raw) or extract_leg_color(it.product_raw),
                "qty": _to_int(it.qty_raw),
                "ship_raw": it.ship_raw,
            }
        )

    orders_df = pd.DataFrame(list(order_rows.values()))
    items_df = pd.DataFrame(item_out)

    # Build order_list summary from items
    if len(orders_df) and len(items_df):
        def _s(v: Any) -> str:
            if v is None:
                return ""
            try:
                # pandas NaN
                if pd.isna(v):
                    return ""
            except Exception:
                pass
            return str(v).strip()

        def _item_line(r: pd.Series) -> str:
            name = _s(r.get("product_canonical")) or _s(r.get("product_raw")) or _s(r.get("spec_raw")) or _s(r.get("note_raw"))
            spec = _s(r.get("spec_raw"))
            size = _s(r.get("size"))
            shelf = _s(r.get("shelf_color"))
            leg = _s(r.get("leg_color"))
            qty = r.get("qty")
            qty_s = "" if qty is None or (hasattr(pd, "isna") and pd.isna(qty)) else str(int(qty))
            parts = [
                name,
                f"규격:{spec}" if spec else "규격:",
                f"사이즈:{size}" if size else "사이즈:",
                f"책장색상:{shelf}" if shelf else "책장색상:",
                f"다리색상:{leg}" if leg else "다리색상:",
                f"개수:{qty_s}" if qty_s else "개수:",
            ]
            return " | ".join(parts).strip(" |")

        tmp = items_df.copy()
        tmp["_line"] = tmp.apply(_item_line, axis=1)
        summary = tmp.groupby("order_id")["_line"].apply(lambda xs: "\n".join([x for x in xs if x])).reset_index()
        # Avoid duplicate order_list column: replace it with computed summary.
        if "order_list" in orders_df.columns:
            orders_df = orders_df.drop(columns=["order_list"])
        orders_df = orders_df.merge(summary, on="order_id", how="left").rename(columns={"_line": "order_list"})

    if len(orders_df):
        orders_df = apply_drawing_ref_status_from_sheet(orders_df, items_df)

    # Write unknown product report as a side effect (easy for user to update aliases)
    if unknown_products:
        unk = (
            pd.DataFrame([{"product_key": k, "count": v} for k, v in unknown_products.items()])
            .sort_values(["count", "product_key"], ascending=[False, True])
            .reset_index(drop=True)
        )
        unk.to_csv("unknown_products.csv", index=False, encoding="utf-8-sig")

    # Empty inputs: keep stable SQLite schema (column names) for the dashboard.
    if len(orders_df) == 0:
        orders_df = pd.DataFrame(
            columns=[
                "order_id",
                "source_file",
                "group_no",
                "group_start_row",
                "deadline_raw",
                "order_date_raw",
                "purchase_date",
                "receiver_name",
                "address",
                "phone",
                "delivery_request",
                "order_list",
                "special_issue",
                "status",
                "shipped_at",
                "created_at",
                "phone_norm",
                "party_key",
            ]
        )
    else:
        orders_df = orders_df.sort_values(["source_file", "group_no"], na_position="last")
        orders_df = assign_recipient_ids(orders_df)

    if len(items_df) == 0:
        items_df = pd.DataFrame(
            columns=[
                "order_id",
                "source_file",
                "row_idx",
                "product_raw",
                "product_key",
                "product_canonical",
                "suggested_canonical",
                "suggestion_score",
                "spec_raw",
                "note_raw",
                "size",
                "shelf_color",
                "leg_color",
                "qty",
                "ship_raw",
            ]
        )
    else:
        items_df = items_df.sort_values(["source_file", "order_id", "row_idx"], na_position="last")

    return orders_df, items_df


ORDER_EDITABLE_COLS = {
    "receiver_name",
    "address",
    "phone",
    "delivery_request",
    "order_list",
    "special_issue",
    "status",
    "shipped_at",
    # 대시보드에서 기록하는 마감 시각 — 재수집 시 incoming에 없으므로 반드시 보존
    "closed_at",
}

# 재수집 시 엑셀 파싱 기본값이 아닌, DB·대시보드에 이미 쌓인 값을 우선한다.
ORDER_USER_PRESERVED_COLS: frozenset[str] = frozenset(
    {"status", "shipped_at", "closed_at", "special_issue"},
)


def _merge_cell_meaningful(v: object) -> bool:
    if v is None:
        return False
    try:
        if pd.isna(v):
            return False
    except Exception:
        pass
    if isinstance(v, str) and not v.strip():
        return False
    return True


def _merge_orders_preserving_edits(existing: pd.DataFrame, incoming: pd.DataFrame) -> pd.DataFrame:
    """
    incoming(새 파싱)을 베이스로 하되, 기존 주문에 대해서는
    - 워크플로·수기 필드(status, shipped_at, closed_at, special_issue)는 DB 값이 있으면 유지
    - 연락·목록 등은 새 값이 비어 있을 때만 기존으로 보완
    """
    if existing is None or len(existing) == 0:
        return incoming
    if "order_id" not in existing.columns:
        return incoming

    ex = existing.copy()
    inc = incoming.copy()
    ex = ex.set_index("order_id", drop=False)
    inc = inc.set_index("order_id", drop=False)

    # Ensure editable columns exist in incoming so they can be preserved
    for col in ORDER_EDITABLE_COLS:
        if col not in inc.columns:
            inc[col] = None

    merged = inc.copy()
    for col in ORDER_EDITABLE_COLS:
        if col not in ex.columns:
            continue
        if col in ORDER_USER_PRESERVED_COLS:
            if col == "status":
                # 기본은 DB 값 우선이나, 엑셀에서 '도면참조'가 잡히면 기존 **접수**만 신규 값으로 승격
                take_ex = ex[col].map(_merge_cell_meaningful).reindex(merged.index).fillna(False)
                exs = ex[col].astype(str).str.strip().reindex(merged.index).fillna("")
                incs = merged[col].astype(str).str.strip()
                upgrade = take_ex & exs.eq("접수") & incs.eq("도면참조")
                merged[col] = merged[col].where(~take_ex | upgrade, ex[col].reindex(merged.index))
            else:
                take_ex = ex[col].map(_merge_cell_meaningful).fillna(False)
                merged[col] = ex[col].where(take_ex, merged[col])
        else:
            take_inc = merged[col].map(_merge_cell_meaningful).fillna(False)
            merged[col] = merged[col].where(take_inc, ex[col])

    # Also keep created_at if it already existed (so "오늘 접수" 기준이 안정적)
    if "created_at" in ex.columns:
        merged["created_at"] = merged["created_at"].where(merged["created_at"].notna(), ex["created_at"])
        merged["created_at"] = merged["created_at"].where(~merged.index.isin(ex.index), ex["created_at"])

    # NOTE: Do not keep orders that disappeared from incoming.
    # Keeping them can create "ghost" orders when parsing rules/order_id change.

    merged = merged.reset_index(drop=True)
    return merged


def write_sqlite(db_path: str, orders_df: pd.DataFrame, items_df: pd.DataFrame) -> None:
    con = sqlite3.connect(db_path)
    try:
        # Preserve user-entered columns in orders table on re-ingest.
        try:
            existing_orders = pd.read_sql_query("select * from orders", con)
        except Exception:
            existing_orders = pd.DataFrame()

        merged_orders = _merge_orders_preserving_edits(existing_orders, orders_df)
        merged_orders = assign_recipient_ids(merged_orders)
        merged_orders.to_sql("orders", con, if_exists="replace", index=False)

        # Items are derived from xlsx; safe to replace.
        items_df.to_sql("items", con, if_exists="replace", index=False)
    finally:
        con.close()


def iter_xlsx_files(input_dir: str) -> list[str]:
    out: list[str] = []
    if not os.path.isdir(input_dir):
        return out
    for root, _dirs, files in os.walk(input_dir):
        for name in files:
            if name.startswith("~$"):
                continue
            if name.lower().endswith(".xlsx"):
                out.append(os.path.join(root, name))
    return sorted(out)


_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


def run_ingest(
    db_path: str,
    *,
    input_dir: str = "order_list",
    aliases_path: str = "product_aliases.yml",
) -> tuple[bool, str]:
    """`order_list` 등의 .xlsx를 읽어 SQLite에 반영. cwd와 무관하게 이 저장소 폴더(스크립트 위치)를 기준으로 한다.

    대시보드에서 호출할 때 터미널 ingest와 동일한 결과를 내도록 한다.
    """
    root = _SCRIPT_DIR

    def _to_abs(p: str) -> str:
        p = (p or "").strip()
        if not p:
            return ""
        return os.path.normpath(p if os.path.isabs(p) else os.path.join(root, p))

    inp = _to_abs(input_dir) or os.path.join(root, "order_list")
    db_abs = _to_abs(db_path) or os.path.join(root, "mutomo.sqlite")
    ali = _to_abs(aliases_path) or os.path.join(root, "product_aliases.yml")

    if not os.path.isdir(inp):
        default_order_list = os.path.normpath(os.path.join(root, "order_list"))
        if os.path.normpath(inp) == default_order_list:
            try:
                os.makedirs(inp, exist_ok=True)
            except OSError as e:
                return False, f"order_list 폴더를 만들 수 없습니다: {e}"
        else:
            return False, f"입력 폴더가 없습니다: `{inp}`"

    old_cwd = os.getcwd()
    try:
        os.chdir(root)
        paths = iter_xlsx_files(inp)
        if not paths:
            orders_df, items_df = build_frames([], ali)
            write_sqlite(db_abs, orders_df, items_df)
            return (
                True,
                f"xlsx 없음 — 빈 orders/items 반영: 주문 {len(orders_df)}건 (`{inp}` → `{db_abs}`)",
            )
        all_rows: list[ItemRow] = []
        for p in paths:
            all_rows.extend(_parse_excel(p))
        orders_df, items_df = build_frames(all_rows, ali)
        write_sqlite(db_abs, orders_df, items_df)
        extra = ""
        if os.path.isfile(os.path.join(root, "unknown_products.csv")):
            extra = " (unknown_products.csv 갱신)"
        return True, f"수집 완료: xlsx {len(paths)}개 → 주문 {len(orders_df)}건, 품목 {len(items_df)}줄{extra}"
    except Exception as e:
        return False, f"수집 실패: {e!s}"
    finally:
        try:
            os.chdir(old_cwd)
        except Exception:
            pass


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--input-dir",
        default="order_list",
        help=".xlsx 수집 폴더 (기본: 프로젝트 기준 order_list)",
    )
    ap.add_argument("--aliases", default="product_aliases.yml", help="yaml mapping of canonical -> aliases")
    ap.add_argument("--db", default="mutomo.sqlite", help="output sqlite database path")
    args = ap.parse_args()

    ok, msg = run_ingest(args.db, input_dir=args.input_dir, aliases_path=args.aliases)
    print(msg)
    if not ok:
        raise SystemExit(msg)
    if os.path.exists(os.path.join(_SCRIPT_DIR, "unknown_products.csv")):
        print("Also wrote unknown product keys to unknown_products.csv")


if __name__ == "__main__":
    main()

