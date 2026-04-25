from __future__ import annotations

import datetime as dt
import difflib
import io
import os
import re
import sqlite3
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import pandas as pd
import streamlit as st
import streamlit.components.v1 as st_components
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ingest_xlsx import (
    apply_drawing_ref_status_from_sheet,
    classify_ship_raw,
    infer_settlement_ship_series,
    order_sheet_blob_for_drawing_ref,
    run_ingest,
    sheet_contains_drawing_ref_keyword,
    shelf_color_from_note_raw_cell,
)
from pricing import format_won, line_unit_prices, load_unit_prices, lookup_line_price
from recipient_identity import assign_recipient_ids
from sales_period_agg import summarize_sales_period


def _note_icons(note_text: str) -> str:
    """Derive quick icons from attention-note text."""
    t = (note_text or "").strip()
    if not t:
        return ""

    icons: list[str] = []
    # Cancellation / cancel request
    if any(k in t for k in ["취소", "취소요청", "취소 요청", "주문취소", "주문 취소"]):
        icons.append("⛔")
    # Product change (not necessarily color-only)
    if "제품변경" in t or "제품 변경" in t:
        icons.append("🔄")
    # Color change only (색상/컬러 변경)
    if any(k in t for k in ["색상변경", "색상 변경", "컬러변경", "컬러 변경", "색깔변경", "색깔 변경"]):
        icons.append("🔄")
    # Addition
    if any(k in t for k in ["추가", "추가됨", "추가했습니다"]):
        icons.append("➕")
    # Color designation (pin)
    # Common vendor phrasing: "컬러지정", "색상지정"
    if ("컬러지정" in t) or ("색상지정" in t) or ("색깔지정" in t):
        icons.append("📌")
    elif any(k in t for k in ["컬러", "색상", "색깔"]) and ("지정" in t) and ("변경" not in t):
        icons.append("📌")

    # Remove duplicates while preserving order
    out: list[str] = []
    for ic in icons:
        if ic not in out:
            out.append(ic)
    return " ".join(out) if out else ""


def _status_lead_icon(status: str) -> str:
    """Exactly one status emoji before the name (empty for 접수/기타)."""
    s = (status or "").strip()
    if s == "출고":
        return "🚚 "
    if s == "클레임":
        return "⚠️ "
    if s == "마감":
        return "💰 "
    if s == "납품취소":
        return "⛔ "
    return ""


def _truncate_display_name(name: str, max_chars: int = 8) -> str:
    n = (name or "").strip()
    if len(n) <= max_chars:
        return n
    if max_chars < 2:
        return n[:max_chars]
    return n[: max_chars - 1] + "…"


def _attention_note_str(att_v: object) -> tuple[str, bool]:
    if att_v is None:
        return "", False
    try:
        if pd.isna(att_v):
            return "", False
    except Exception:
        pass
    s = str(att_v).strip()
    return s, bool(s)


def _fold_ws(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def _strip_order_list_overlap(text_v: object, order_list_v: object) -> str:
    """특이사항/수기 메모에서 DB `order_list`(품목 요약)와 동일한 줄·블록을 제거한다.

    엑셀 A열에 주문목록을 붙여 넣은 경우 피킹·출고_주문에서 특이사항과 품목이 이중으로 보이는 것을 막는다.
    """
    body, has = _attention_note_str(text_v)
    if not has:
        return ""
    ol = str(order_list_v or "").strip()
    if not ol:
        return body

    if ol in body:
        body = body.replace(ol, "").strip()
        if not body:
            return ""

    ol_lines = {_fold_ws(x) for x in ol.splitlines() if _fold_ws(x)}
    if not ol_lines:
        return body

    kept: list[str] = []
    for line in body.splitlines():
        if _fold_ws(line) in ol_lines:
            continue
        kept.append(line)
    body = "\n".join(kept).strip()

    # 한 줄에 요약 전체가 이어 붙은 경우
    ol2 = ol.strip()
    if len(ol2) > 40 and ol2 in body:
        body = body.replace(ol2, "").strip()

    return body


def _attention_note_export_for_order(att_v: object, order_list_v: object) -> str:
    cleaned = _strip_order_list_overlap(att_v, order_list_v)
    return _attention_note_with_icon_export(cleaned or None)


def _combined_attention_for_icons(attention_note_val: object, special_issue_val: object) -> object:
    """이름 옆 아이콘용: 엑셀 특이사항(자동) + 수기 이슈를 함께 본다(본문 열과는 별개)."""
    parts: list[str] = []
    for v in (attention_note_val, special_issue_val):
        if v is None:
            continue
        try:
            if pd.isna(v):
                continue
        except Exception:
            pass
        s = str(v).strip()
        if not s or s.lower() in ("nan", "none", "-", "#n/a"):
            continue
        parts.append(s)
    if not parts:
        return None
    return "\n".join(parts)


def _trailing_icon_segment(has_attention: bool, note_text: str, max_icons: int = 3) -> str:
    """Up to max_icons after the name: 🟥 first if present, then note-derived icons."""
    seq: list[str] = []
    if has_attention:
        seq.append("🟥")
    for tok in _note_icons(note_text).split():
        if tok:
            seq.append(tok)
    seen: set[str] = set()
    out: list[str] = []
    for t in seq:
        if t not in seen:
            seen.add(t)
            out.append(t)
        if len(out) >= max_icons:
            break
    return (" " + " ".join(out)) if out else ""


def _drawing_ref_sheet_blob(row: pd.Series, items_df: pd.DataFrame | None) -> str:
    """배송요청·품목줄 등 주문서 전역에서 도면참조/도면참고 탐지용(ingest와 동일 범위)."""
    if items_df is not None and len(items_df) and "order_id" in items_df.columns and "order_id" in row.index:
        return order_sheet_blob_for_drawing_ref(row, items_df)
    parts: list[str] = []
    for key in ("receiver_name", "delivery_request", "attention_note", "order_list", "special_issue", "deadline_raw"):
        if key in row.index:
            parts.append(str(row.get(key) or ""))
    return "\n".join(parts)


def _is_drawing_ref_order_row(row: pd.Series, items_df: pd.DataFrame | None) -> bool:
    try:
        return sheet_contains_drawing_ref_keyword(_drawing_ref_sheet_blob(row, items_df))
    except Exception:
        return False


def _compact_name_display(
    status: str, receiver_name: str, attention_note_val: object, *, sheet_blob: str = ""
) -> str:
    att, has_att = _attention_note_str(attention_note_val)
    lead = _status_lead_icon(status)
    nick = _truncate_display_name(receiver_name, 8)
    # 도면참조(구 주문제작): 이름 뒤 📐 — 본문에 도면참조·도면참고(엑셀 표기)가 있으면 표시
    ss = (status or "").strip()
    draw_after = ""
    if ss in ("도면참조", "주문제작") or sheet_contains_drawing_ref_keyword(sheet_blob):
        draw_after = " 📐"
    tail = _trailing_icon_segment(has_att, att, 3)
    return f"{lead}{nick}{draw_after}{tail}".strip()


def _attention_note_with_icon_export(att_val: object) -> str:
    """특이사항(자동) 본문 앞에 목록·엑셀과 동일한 아이콘 줄(🟥·🔄 등)을 붙인다."""
    att, has_att = _attention_note_str(att_val)
    icons = _trailing_icon_segment(has_att, att, 5).strip()
    body = att.strip()
    if icons and body:
        return f"{icons}\n\n{body}"
    if body:
        return body
    return icons


# 자동 백업: 로컬 하루 최대 4번(6시간 구간당 1회), 파일명 시각 기준 keep_days 초과분 삭제
BACKUP_RETENTION_DAYS = 30
BACKUPS_PER_LOCAL_DAY = 4


def _backup_db_basename(db_path: str) -> str:
    return os.path.splitext(os.path.basename(db_path))[0] or "mutomo"


def _backup_time_slot(now: dt.datetime | None = None) -> tuple[str, int]:
    """로컬 날짜 YYYYMMDD와 0..BACKUPS_PER_LOCAL_DAY-1 구간(각 24/n 시간)."""
    n = now or dt.datetime.now()
    span = max(1, 24 // BACKUPS_PER_LOCAL_DAY)
    return n.strftime("%Y%m%d"), n.hour // span


def _sqlite_slot_already_has_backup(
    db_path: str, backup_dir: str = "backups", now: dt.datetime | None = None
) -> bool:
    """같은 로컬일·같은 시간 구간에 이미 백업 sqlite가 있으면 True."""
    n = now or dt.datetime.now()
    day_s, slot = _backup_time_slot(n)
    span = max(1, 24 // BACKUPS_PER_LOCAL_DAY)
    h0, h1 = slot * span, (slot + 1) * span
    base = _backup_db_basename(db_path)
    prefix = f"{base}_{day_s}_"
    try:
        names = os.listdir(backup_dir)
    except OSError:
        return False
    for name in names:
        if not name.startswith(prefix) or not name.endswith(".sqlite"):
            continue
        stamp = name[len(base) + 1 :].replace(".sqlite", "")
        try:
            d = dt.datetime.strptime(stamp, "%Y%m%d_%H%M%S")
        except Exception:
            continue
        if h0 <= d.hour < h1:
            return True
    return False


def _prune_sqlite_backups(db_path: str, backup_dir: str = "backups", keep_days: int = BACKUP_RETENTION_DAYS) -> None:
    """파일명 타임스탬프 기준 keep_days보다 오래된 동일-베이스 백업을 삭제(자동 백업과 CLI 공통 규칙)."""
    base = _backup_db_basename(db_path)
    try:
        cutoff = dt.datetime.now() - dt.timedelta(days=keep_days)
        for name in os.listdir(backup_dir):
            if not name.startswith(base + "_") or not name.endswith(".sqlite"):
                continue
            stamp = name[len(base) + 1 :].replace(".sqlite", "")
            try:
                d = dt.datetime.strptime(stamp, "%Y%m%d_%H%M%S")
            except Exception:
                continue
            if d < cutoff:
                try:
                    os.remove(os.path.join(backup_dir, name))
                except Exception:
                    pass
    except Exception:
        pass


def _backup_sqlite(db_path: str, backup_dir: str = "backups", keep_days: int = BACKUP_RETENTION_DAYS) -> str | None:
    """Create a safe SQLite backup file and return its path.

    Uses sqlite3.Connection.backup() so it works even while the DB is in use.
    """
    try:
        os.makedirs(backup_dir, exist_ok=True)
    except Exception:
        return None

    if not db_path or not os.path.exists(db_path):
        return None

    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    base = _backup_db_basename(db_path)
    out_path = os.path.join(backup_dir, f"{base}_{ts}.sqlite")

    try:
        src = sqlite3.connect(db_path)
        try:
            dst = sqlite3.connect(out_path)
            try:
                src.backup(dst)
            finally:
                dst.close()
        finally:
            src.close()
    except Exception:
        try:
            if os.path.exists(out_path):
                os.remove(out_path)
        except Exception:
            pass
        return None

    _prune_sqlite_backups(db_path, backup_dir, keep_days)
    return out_path


@st.cache_data
def _price_map_cached(price_path: str, price_mtime: float) -> tuple[dict, tuple[str, ...]]:
    return load_unit_prices(price_path if os.path.isfile(price_path) else None)


@st.cache_data
def load_tables(db_path: str, db_mtime: float, db_size: int, reload_nonce: int = 0) -> tuple[pd.DataFrame, pd.DataFrame]:
    """db_mtime/db_size/reload_nonce는 캐시 키에 포함(언더스코어 접두 인자는 키에서 빠짐).

    SQLite 파일이 바뀌면 mtime·size가 달라져 자동으로 다시 읽고, 'DB 다시 읽기'는 nonce로 강제 무효화합니다.
    """
    _ = (db_mtime, db_size, reload_nonce)
    con = sqlite3.connect(db_path)
    try:
        orders = pd.read_sql_query("select * from orders", con)
        items = pd.read_sql_query("select * from items", con)
    finally:
        con.close()
    if "status" in orders.columns:
        s0 = orders["status"].astype(str).str.strip()
        orders = orders.copy()
        orders["status"] = s0.mask(s0 == "주문제작", "도면참조")
    # 엑셀 본문에 도면참조·도면참고가 있고 status가 접수인 주문 → 도면참조로 승격(재수집 시에도 동일)
    orders = apply_drawing_ref_status_from_sheet(orders, items)
    return orders, items


def _db_stat_for_cache(db_path: str) -> tuple[float, int]:
    try:
        stt = os.stat(db_path)
        return (float(stt.st_mtime), int(stt.st_size))
    except OSError:
        return (0.0, 0)


def _db_ready(db_path: str) -> tuple[bool, str]:
    """Return (ok, reason) — dashboard needs orders + items tables from ingest."""
    path = (db_path or "").strip()
    if not path:
        return False, "DB 경로가 비어 있습니다. 사이드바 **설정**에서 경로를 지정하세요."
    if not os.path.exists(path):
        return False, f"DB 파일이 없습니다: `{path}`"
    con = sqlite3.connect(path)
    try:
        cur = con.cursor()
        cur.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name IN ('orders', 'items')"
        )
        found = {row[0] for row in cur.fetchall()}
    finally:
        con.close()
    missing = [t for t in ("orders", "items") if t not in found]
    if missing:
        return (
            False,
            f"SQLite에 필요한 테이블이 없습니다: {', '.join(missing)}. "
            "엑셀 수집(`ingest_xlsx.py`)으로 DB를 만든 뒤 다시 실행하세요.",
        )
    return True, ""


def _migrate_orders_schema(db_path: str) -> None:
    """orders에 shipped_at·closed_at·phone_norm·party_key 컬럼 보장 및 party 컬럼 백필."""
    con = sqlite3.connect(db_path)
    try:
        cur = con.cursor()
        cur.execute("PRAGMA table_info(orders)")
        cols = {row[1] for row in cur.fetchall()}
        if "shipped_at" not in cols:
            cur.execute("ALTER TABLE orders ADD COLUMN shipped_at TEXT")
        if "closed_at" not in cols:
            cur.execute("ALTER TABLE orders ADD COLUMN closed_at TEXT")
        if "phone_norm" not in cols:
            cur.execute("ALTER TABLE orders ADD COLUMN phone_norm TEXT")
        if "party_key" not in cols:
            cur.execute("ALTER TABLE orders ADD COLUMN party_key TEXT")
        con.commit()

        cur.execute("SELECT COUNT(*) FROM orders")
        n_orders = int(cur.fetchone()[0] or 0)
        if n_orders:
            cur.execute(
                "SELECT COUNT(*) FROM orders WHERE party_key IS NULL OR TRIM(COALESCE(party_key,'')) = ''"
            )
            need_fill = int(cur.fetchone()[0] or 0) > 0
            if need_fill:
                df = pd.read_sql_query("SELECT * FROM orders", con)
                df = assign_recipient_ids(df)
                cur.executemany(
                    "UPDATE orders SET phone_norm = ?, party_key = ? WHERE order_id = ?",
                    list(
                        zip(
                            df["phone_norm"].astype(str).tolist(),
                            df["party_key"].astype(str).tolist(),
                            df["order_id"].astype(str).tolist(),
                            strict=False,
                        )
                    ),
                )
                con.commit()
    finally:
        con.close()


def _to_date_series(s: pd.Series) -> pd.Series:
    # created_at is ISO string; coerce anything else safely
    return pd.to_datetime(s, errors="coerce").dt.date


try:
    _SEOUL = ZoneInfo("Asia/Seoul")
except ZoneInfoNotFoundError:
    # Windows 환경에서 tz database가 없으면 ZoneInfo가 실패할 수 있음 → 고정 KST(+09:00)로 폴백
    _SEOUL = dt.timezone(dt.timedelta(hours=9), name="KST")


def _shipped_at_calendar_date_series(shipped_at: pd.Series) -> pd.Series:
    """출고일 집계·엑셀 필터용: shipped_at → 대한민국 달력 날짜.

    - timezone-aware(예: ...+00:00)는 **Asia/Seoul**로 바꾼 뒤 날짜를 씁니다.
      (전날 UTC 오후 시각이 한국에서는 다음 날 자정인 경우 UTC `.date()`만 쓰면 하루 밀림)
    - naïve ISO(대시보드에서 `datetime.now().isoformat()`으로 넣은 값 등)는 **그 벽시계 그대로** 날짜를 씁니다.
    """
    out: list[dt.date | None] = []
    for v in shipped_at.tolist():
        if v is None:
            out.append(None)
            continue
        try:
            if isinstance(v, float) and pd.isna(v):
                out.append(None)
                continue
        except Exception:
            pass
        t = pd.to_datetime(v, errors="coerce")
        if pd.isna(t):
            out.append(None)
            continue
        ts = pd.Timestamp(t)
        try:
            if ts.tzinfo is None:
                ts2 = ts.tz_localize(_SEOUL, ambiguous="NaT", nonexistent="NaT")
            else:
                ts2 = ts.tz_convert(_SEOUL)
        except Exception:
            ts2 = ts
        if pd.isna(ts2):
            out.append(None)
        else:
            out.append(pd.Timestamp(ts2).date())
    return pd.Series(out, index=shipped_at.index, dtype=object)


def _ensure_settlements_table(db_path: str) -> None:
    """Create settlements table if missing (idempotent)."""
    con = sqlite3.connect(db_path)
    try:
        cur = con.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS settlements (
              period_start TEXT NOT NULL,
              period_end   TEXT NOT NULL,
              basis        TEXT NOT NULL,
              expected_sale REAL DEFAULT 0,
              expected_gj   REAL DEFAULT 0,
              paid_amount   REAL DEFAULT 0,
              note          TEXT,
              updated_at    TEXT,
              PRIMARY KEY (period_start, period_end, basis)
            )
            """
        )
        con.commit()
    finally:
        con.close()


def _ensure_order_amount_overrides_table(db_path: str) -> None:
    """Manual per-order amount overrides (for custom/drawing reference orders)."""
    con = sqlite3.connect(db_path)
    try:
        cur = con.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS order_amount_overrides (
              order_id     TEXT PRIMARY KEY,
              sale_amount  REAL DEFAULT 0,
              gj_amount    REAL DEFAULT 0,
              note         TEXT,
              updated_at   TEXT
            )
            """
        )
        con.commit()
    finally:
        con.close()


def _load_order_amount_overrides(db_path: str, order_ids: set[str]) -> dict[str, tuple[float, float]]:
    """Return order_id -> (sale_amount, gj_amount) for orders with overrides."""
    out: dict[str, tuple[float, float]] = {}
    if not order_ids:
        return out
    _ensure_order_amount_overrides_table(db_path)
    con = sqlite3.connect(db_path)
    try:
        cur = con.cursor()
        ids = [str(x) for x in order_ids if str(x)]
        # SQLite variable limit is high enough for our use; chunk defensively anyway.
        chunk = 450
        for i in range(0, len(ids), chunk):
            sub = ids[i : i + chunk]
            qmarks = ",".join(["?"] * len(sub))
            cur.execute(
                f"SELECT order_id, sale_amount, gj_amount FROM order_amount_overrides WHERE order_id IN ({qmarks})",
                sub,
            )
            for oid, sale, gj in cur.fetchall():
                try:
                    out[str(oid)] = (float(sale or 0.0), float(gj or 0.0))
                except Exception:
                    out[str(oid)] = (0.0, 0.0)
    finally:
        con.close()
    return out


def _upsert_order_amount_override(db_path: str, order_id: str, sale_amount: float, gj_amount: float, note: str = "") -> None:
    _ensure_order_amount_overrides_table(db_path)
    con = sqlite3.connect(db_path)
    try:
        cur = con.cursor()
        cur.execute(
            """
            INSERT INTO order_amount_overrides(order_id, sale_amount, gj_amount, note, updated_at)
            VALUES(?, ?, ?, ?, ?)
            ON CONFLICT(order_id) DO UPDATE SET
              sale_amount=excluded.sale_amount,
              gj_amount=excluded.gj_amount,
              note=excluded.note,
              updated_at=excluded.updated_at
            """,
            (
                str(order_id),
                float(sale_amount or 0.0),
                float(gj_amount or 0.0),
                str(note or ""),
                dt.datetime.now().isoformat(timespec="seconds"),
            ),
        )
        con.commit()
    finally:
        con.close()


def _calc_amounts_for_orders(
    items_df: pd.DataFrame, order_ids: set[str], price_map: dict, overrides: dict[str, tuple[float, float]] | None = None
) -> tuple[float, float, int, int]:
    """Return (sale_amount, gwangjin_amount, n_unpriced_lines, unpriced_qty) for given orders."""
    if not order_ids:
        return 0.0, 0.0, 0, 0
    ov = overrides or {}
    ov_keys = set(ov.keys())
    remaining = set(order_ids) - ov_keys
    base_sale = float(sum(ov.get(oid, (0.0, 0.0))[0] for oid in ov_keys))
    base_gj = float(sum(ov.get(oid, (0.0, 0.0))[1] for oid in ov_keys))
    if items_df is None or len(items_df) == 0 or "order_id" not in items_df.columns or not remaining:
        return base_sale, base_gj, 0, 0
    it = items_df[items_df["order_id"].astype(str).isin(remaining)].copy()
    if len(it) == 0:
        return base_sale, base_gj, 0, 0
    q = pd.to_numeric(it.get("qty"), errors="coerce").fillna(0).astype(int).clip(lower=0)
    it["_q"] = q
    sale_amount = 0.0
    gj_amount = 0.0
    n_unpriced = 0
    unpriced_qty = 0
    for _, r in it.iterrows():
        qty_i = int(r.get("_q") or 0)
        pu = line_unit_prices(r, price_map) if price_map else None
        if pu is None:
            n_unpriced += 1
            unpriced_qty += qty_i
            continue
        try:
            sale_amount += float(pu[0] or 0.0) * qty_i
        except (TypeError, ValueError):
            pass
        try:
            gj_amount += float(pu[1] or 0.0) * qty_i
        except (TypeError, ValueError):
            pass
    return float(base_sale + sale_amount), float(base_gj + gj_amount), int(n_unpriced), int(unpriced_qty)


def _calc_order_amount_map(
    items_df: pd.DataFrame,
    order_ids: set[str],
    price_map: dict,
    overrides: dict[str, tuple[float, float]] | None = None,
) -> dict[str, tuple[float, float]]:
    """order_id -> (sale_amount, gwangjin_amount) using 단가표. Missing prices contribute 0."""
    out: dict[str, tuple[float, float]] = {}
    if not order_ids:
        return out
    ov = overrides or {}
    for oid, (sa, gj) in ov.items():
        if oid in order_ids:
            out[str(oid)] = (float(sa or 0.0), float(gj or 0.0))
    remaining = set(order_ids) - set(out.keys())
    if not remaining or items_df is None or len(items_df) == 0 or "order_id" not in items_df.columns:
        return out
    it = items_df[items_df["order_id"].astype(str).isin(remaining)].copy()
    if len(it) == 0:
        return out
    it["_oid"] = it["order_id"].astype(str)
    it["_q"] = pd.to_numeric(it.get("qty"), errors="coerce").fillna(0).astype(int).clip(lower=0)
    acc_sale: dict[str, float] = {}
    acc_gj: dict[str, float] = {}
    for _, r in it.iterrows():
        oid = str(r.get("_oid") or "").strip()
        if not oid:
            continue
        q = int(r.get("_q") or 0)
        pu = line_unit_prices(r, price_map) if price_map else None
        if pu is None or q <= 0:
            continue
        try:
            acc_sale[oid] = acc_sale.get(oid, 0.0) + float(pu[0] or 0.0) * q
        except (TypeError, ValueError):
            pass
        try:
            acc_gj[oid] = acc_gj.get(oid, 0.0) + float(pu[1] or 0.0) * q
        except (TypeError, ValueError):
            pass
    for oid in remaining:
        out[str(oid)] = (float(acc_sale.get(str(oid), 0.0)), float(acc_gj.get(str(oid), 0.0)))
    return out


def _ship_bucket_for_order(items_all: pd.DataFrame, order_id: str) -> str:
    """택배/직접/혼합 (품목 배송란 기준)."""
    ht, hd = _order_picking_sheet_hits(items_all, order_id)
    if ht and hd:
        return "혼합"
    if hd:
        return "직접"
    return "택배"


@st.cache_data
def _today_shipped_excel_cached(export_version: str, shipped_key: str, orders_all: pd.DataFrame, items_all: pd.DataFrame) -> bytes:
    # shipped_key forces cache bust when today's shipped set changes
    # shipped_key includes date selection
    shipped_date = dt.date.fromisoformat(shipped_key.split("|", 1)[0])
    return _build_shipped_excel_bytes(orders_all, items_all, shipped_date=shipped_date)


def _excel_line_pick_bucket(raw: object) -> str:
    """엑셀 품목 배송란 기준으로 피킹 시트(택배/직접) 분류. 미기재·판별불가는 택배 쪽에 둡니다."""
    if classify_ship_raw(raw) == "직접배송":
        return "직접배송"
    return "택배"


def _ship_display_label(canonical: str) -> str:
    """엑셀과 동일하게 표기: 직접배송 → 직접."""
    c = (canonical or "").strip()
    if c == "직접배송":
        return "직접"
    if c == "택배":
        return "택배"
    return c or "택배"


def _excel_pick_bucket_series(base: pd.DataFrame) -> pd.Series:
    """주문 단위 품목 행마다 택배/직접배송 버킷. 배송란 비어 있으면 같은 주문 위쪽 행 값(ffill)."""
    if "ship_raw" not in base.columns:
        return pd.Series(_excel_line_pick_bucket(None), index=base.index, dtype=object)
    work = base.copy()
    if "row_idx" in work.columns:
        work = work.sort_values("row_idx")

    def _is_blank(v: object) -> bool:
        if v is None:
            return True
        try:
            if pd.isna(v):
                return True
        except Exception:
            pass
        return isinstance(v, str) and not str(v).strip()

    sr = work["ship_raw"].astype(object)
    filled = sr.mask(sr.map(_is_blank), pd.NA).ffill()
    work["_pick_bucket"] = filled.map(_excel_line_pick_bucket)
    return work.loc[base.index, "_pick_bucket"]


def _excel_ship_for_picking_row(order_id: object, items_df: pd.DataFrame, pick_kind: str) -> str:
    """피킹 시트에 실린 품목 줄만으로 다수결(주문 전체 요약이 아님)."""
    oid = str(order_id or "").strip()
    if not oid or "order_id" not in items_df.columns or "ship_raw" not in items_df.columns:
        return _ship_display_label(pick_kind)
    base = items_df[items_df["order_id"].astype(str) == oid].copy()
    if len(base) == 0:
        return "택배"
    buckets = _excel_pick_bucket_series(base)
    sub = base.loc[buckets == pick_kind]
    if len(sub) == 0:
        return _ship_display_label(pick_kind)
    ser = infer_settlement_ship_series(sub)
    lab = str(ser.iloc[0]) if len(ser) else pick_kind
    return _ship_display_label(lab)


def _items_view_from_item_rows(df: pd.DataFrame) -> pd.DataFrame:
    """이미 주문 1건 분의 품목 행만 담은 데이터프레임 → 표시용 뷰."""
    if df is None or len(df) == 0:
        return pd.DataFrame()

    df = df.copy()

    def _name(r: pd.Series) -> str:
        v = r.get("product_canonical")
        if pd.notna(v) and str(v).strip():
            return str(v).strip()
        v = r.get("product_raw")
        if pd.notna(v) and str(v).strip():
            return str(v).strip()
        return ""

    df["품목"] = df.apply(_name, axis=1)
    # DB에 shelf_color가 비었을 때 note_raw(엑셀 다리색 열 원본)에만 원단명이 있는 경우
    if "shelf_color" in df.columns and "note_raw" in df.columns:
        def _fill_shelf_from_note(r: pd.Series) -> object:
            s = r.get("shelf_color")
            if s is not None and not (isinstance(s, float) and pd.isna(s)) and str(s).strip():
                return s
            fb = shelf_color_from_note_raw_cell(r.get("note_raw"))
            return fb if fb else s

        df["shelf_color"] = df.apply(_fill_shelf_from_note, axis=1)
    # Prefer spec_raw; fallback/append note_raw for options like
    # "(윈터선샤인 1EA, 런던브릭 1EA, ... / 57cm)" that we still want visible.
    if "spec_raw" in df.columns and "note_raw" in df.columns:
        spec = df["spec_raw"].astype(object)
        note = df["note_raw"].astype(object)

        spec_clean = spec.astype(str).fillna("").str.strip()
        note_clean = note.astype(str).fillna("").str.strip()

        # If spec is empty, use note.
        merged = spec.where(spec_clean != "", note)

        # If both exist, append note when it looks like a multi-option description.
        def _should_append(n: str) -> bool:
            n2 = (n or "").strip()
            if not n2:
                return False
            n2l = n2.lower()
            return ("ea" in n2l) or ("," in n2) or ("/" in n2) or ("cm" in n2l)

        out = []
        for s_val, n_val in zip(merged.tolist(), note_clean.tolist(), strict=False):
            s_txt = ("" if s_val is None else str(s_val)).strip()
            n_txt = (n_val or "").strip()
            if s_txt and n_txt and _should_append(n_txt) and n_txt not in s_txt:
                out.append(f"{s_txt} / {n_txt}")
            else:
                out.append(s_txt or n_txt or None)
        df["spec_raw"] = pd.Series(out, index=df.index, dtype=object)
    cols = []
    for c in ["품목", "spec_raw", "size", "shelf_color", "leg_color", "qty", "ship_raw"]:
        if c in df.columns:
            cols.append(c)
    view = df[cols].rename(
        columns={
            "spec_raw": "규격/옵션",
            "size": "사이즈",
            "shelf_color": "책장색상",
            "leg_color": "다리색상",
            "qty": "개수",
            "ship_raw": "배송",
        }
    )
    # Stable ordering
    if "row_idx" in df.columns:
        view = pd.concat([df[["row_idx"]], view], axis=1).sort_values("row_idx").drop(columns=["row_idx"])
    return view


def _items_view(items: pd.DataFrame, order_id: str) -> pd.DataFrame:
    base = items[items["order_id"].astype(str) == str(order_id)].copy() if "order_id" in items.columns else pd.DataFrame()
    return _items_view_from_item_rows(base)


def _items_view_excel_pick_kind(items: pd.DataFrame, order_id: str, pick_kind: str) -> pd.DataFrame:
    """엑셀 ship_raw 기준으로 해당 피킹(택배/직접)에 올릴 품목 줄만 보이게 필터."""
    base = items[items["order_id"].astype(str) == str(order_id)].copy() if "order_id" in items.columns else pd.DataFrame()
    if len(base) == 0:
        return base
    if "ship_raw" not in base.columns:
        return _items_view_from_item_rows(base) if pick_kind == "택배" else pd.DataFrame()
    b = _excel_pick_bucket_series(base)
    sub = base.loc[b == pick_kind].copy()
    return _items_view_from_item_rows(sub)


def _product_catalog_summary(items_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """정규명·미매핑 엑셀명별 (라인 수, 수량 합). `main`에서 쓰는 `items`와 동일 필터 범위."""
    empty_mapped = pd.DataFrame(columns=["정규상품명", "라인수", "수량합"])
    empty_raw = pd.DataFrame(columns=["엑셀상품명", "라인수", "수량합"])
    if items_df is None or len(items_df) == 0:
        return empty_mapped, empty_raw
    df = items_df.copy()
    if "qty" in df.columns:
        df["_qty_num"] = pd.to_numeric(df["qty"], errors="coerce").fillna(0)
    else:
        df["_qty_num"] = 0.0

    if "product_canonical" in df.columns:
        c = df["product_canonical"]
        mask_mapped = c.notna() & c.astype(str).str.strip().ne("")
    else:
        mask_mapped = pd.Series(False, index=df.index)

    mapped = df.loc[mask_mapped]
    if len(mapped):
        g1 = (
            mapped.groupby("product_canonical", dropna=False)
            .agg(라인수=("order_id", "count"), 수량합=("_qty_num", "sum"))
            .reset_index()
            .rename(columns={"product_canonical": "정규상품명"})
            .sort_values(["라인수", "정규상품명"], ascending=[False, True])
        )
    else:
        g1 = empty_mapped.copy()

    mask_un = ~mask_mapped
    if "product_raw" in df.columns:
        pr = df["product_raw"]
        mask_raw = pr.notna() & pr.astype(str).str.strip().ne("")
    else:
        mask_raw = pd.Series(False, index=df.index)
    unmapped = df.loc[mask_un & mask_raw]
    if len(unmapped):
        g2 = (
            unmapped.groupby("product_raw", dropna=False)
            .agg(라인수=("order_id", "count"), 수량합=("_qty_num", "sum"))
            .reset_index()
            .rename(columns={"product_raw": "엑셀상품명"})
            .sort_values(["라인수", "엑셀상품명"], ascending=[False, True])
        )
    else:
        g2 = empty_raw.copy()

    return g1, g2


def _order_picking_sheet_hits(items: pd.DataFrame, order_id: object) -> tuple[bool, bool]:
    """(피킹리스트_택배에 실릴 품목 줄 있음, 피킹리스트_직접에 실릴 품목 줄 있음)."""
    oid = str(order_id or "").strip()
    if not oid:
        return False, False
    ta = _items_view_excel_pick_kind(items, oid, "택배")
    dr = _items_view_excel_pick_kind(items, oid, "직접배송")
    return len(ta) > 0, len(dr) > 0


def _picking_stats(shipped: pd.DataFrame, items_all: pd.DataFrame) -> tuple[int, int, int, int]:
    """(출고 주문 수, 피킹_택배 행 수, 피킹_직접 행 수, 혼합 배송 주문 수). 혼합은 두 피킹 시트에 각 1행."""
    n_u = len(shipped)
    if not n_u:
        return 0, 0, 0, 0
    n_t = n_d = n_m = 0
    for _, o in shipped.iterrows():
        ht, hd = _order_picking_sheet_hits(items_all, o.get("order_id"))
        if ht:
            n_t += 1
        if hd:
            n_d += 1
        if ht and hd:
            n_m += 1
    return n_u, n_t, n_d, n_m


def _items_text(items: pd.DataFrame, order_id: str) -> str:
    view = _items_view(items, order_id)
    if view is None or len(view) == 0:
        return ""

    def _s(v: object) -> str:
        if v is None:
            return ""
        try:
            if pd.isna(v):
                return ""
        except Exception:
            pass
        return str(v).strip().replace("\n", " ")

    lines: list[str] = []
    for _, r in view.iterrows():
        parts = []
        for col in ["품목", "규격/옵션", "사이즈", "책장색상", "다리색상", "개수", "배송"]:
            if col in view.columns:
                val = _s(r.get(col))
                if val:
                    parts.append(f"{col}:{val}")
        if parts:
            lines.append(" | ".join(parts))
    return "\n---\n".join(lines)


def _render_order_detail(container: st.delta_generator.DeltaGenerator, order_row: pd.Series, items: pd.DataFrame) -> None:
    purchase = str(order_row.get("purchase_date") or "").strip()
    name = str(order_row.get("receiver_name") or "").strip()
    phone = str(order_row.get("phone") or "").strip()
    address = str(order_row.get("address") or "").strip()
    req = str(order_row.get("delivery_request") or "").strip()
    att_for_detail = _attention_note_export_for_order(order_row.get("attention_note"), order_row.get("order_list"))
    status = str(order_row.get("status") or "").strip()
    oid = str(order_row.get("order_id") or "").strip()

    # Color-coded header box by status (high contrast)
    def _status_style(s: str) -> tuple[str, str, str, str]:
        # returns (label, bg, border, text)
        if s == "출고":
            return "🚚✅ 출고 완료", "#DFF5E1", "#1B5E20", "#1B5E20"
        if s == "클레임":
            return "⚠️ 클레임", "#FFE4CC", "#E65100", "#8A2E00"
        if s == "마감":
            return "💰✅ 마감(청구)", "#E8F5E9", "#1B5E20", "#1B5E20"
        if s == "접수":
            return "📝⏳ 접수(납품예정)", "#D6ECFF", "#0D47A1", "#0D47A1"
        if s == "도면참조":
            return "📐 도면참조", "#F3E5F5", "#6A1B9A", "#4A148C"
        if s == "납품취소":
            return "⛔ 납품취소", "#FFEBEE", "#B71C1C", "#B71C1C"
        return (s or "상태 없음"), "#ECEFF1", "#607D8B", "#37474F"

    status_label, bg, border, text = _status_style(status)
    container.markdown(
        f"""
<div style="padding:10px 12px; border:2px solid {border}; background:{bg}; border-radius:10px; margin:6px 0 10px 0; box-shadow: 0 1px 4px rgba(0,0,0,0.12);">
  <div style="font-size:12px; font-weight:600; color:#263238; margin-bottom:4px;">상태</div>
  <div style="font-weight:800; color:{text}; font-size:16px;">{status_label}</div>
</div>
""",
        unsafe_allow_html=True,
    )

    header_lines = []
    if purchase:
        header_lines.append(f"- **구매일자**: {purchase}")
    if name:
        header_lines.append(f"- **받는분**: {name}")
    if phone:
        header_lines.append(f"- **전화**: {phone}")
    if address:
        header_lines.append(f"- **주소**: {address}")
    if req:
        header_lines.append(f"- **배송요청**: {req}")
    if att_for_detail:
        header_lines.append(f"- **특이사항(자동)**: {att_for_detail.replace(chr(10), '  \n')}")
    if oid:
        header_lines.append(f"- **ID**: `{oid}`")

    if header_lines:
        container.markdown("\n".join(header_lines))

    view = _items_view(items, oid)
    if view is None or len(view) == 0:
        txt = str(order_row.get("order_list") or "").strip()
        if txt:
            container.markdown("**품목**")
            container.markdown(txt.replace("\n", "  \n"))
        return

    container.markdown("**품목**")
    for i, (_, r) in enumerate(view.iterrows(), start=1):
        def _sv(col: str) -> str:
            if col not in view.columns:
                return ""
            v = r.get(col)
            if v is None:
                return ""
            try:
                if pd.isna(v):
                    return ""
            except Exception:
                pass
            return str(v).strip().replace("\n", " ")

        parts = {
            "품목": _sv("품목"),
            "규격/옵션": _sv("규격/옵션"),
            "사이즈": _sv("사이즈"),
            "책장색상": _sv("책장색상"),
            "다리색상": _sv("다리색상"),
            "개수": _sv("개수"),
            "배송": _sv("배송"),
        }
        title = parts["품목"] or "(품목명 없음)"
        container.markdown(f"**{i}. {title}**")

        # Compact line: 규격+사이즈+책장색상+다리색상+개수+배송 (skip empty)
        compact_parts = []
        if parts.get("규격/옵션"):
            compact_parts.append(parts["규격/옵션"])
        if parts.get("사이즈"):
            compact_parts.append(parts["사이즈"])
        if parts.get("책장색상"):
            compact_parts.append(f"책장:{parts['책장색상']}")
        if parts.get("다리색상"):
            compact_parts.append(f"다리:{parts['다리색상']}")
        if parts.get("개수"):
            compact_parts.append(f"개수:{parts['개수']}")
        if parts.get("배송"):
            compact_parts.append(f"배송:{parts['배송']}")
        if compact_parts:
            container.markdown(" / ".join(compact_parts))
        if i != len(view):
            container.markdown("---")


def _format_picking_worksheet(ws: object) -> None:
    """피킹리스트 시트 공통 인쇄 서식 (택배/직접 동일)."""
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:1"
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="E3F2FD")
    header_font = Font(bold=True)
    thin_side = Side(style="thin", color="B0B0B0")
    header_side = Side(style="medium", color="607D8B")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_border = Border(left=header_side, right=header_side, top=header_side, bottom=header_side)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border

    wrap_cols = {"주소", "배송요청", "특이사항", "품목"}
    header_map = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for name, col_idx in header_map.items():
        if name == "No":
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 5
        elif name in ("받는분",):
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 12
        elif name == "배송(엑셀)":
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 11
        elif name in ("전화",):
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 14
        elif name == "주소":
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 40
        elif name == "배송요청":
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 24
        elif name == "특이사항":
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 44
        elif name == "품목":
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 52

    item_col_idx = header_map.get("품목")
    notes_col_idx = header_map.get("특이사항")
    item_col_width = 52
    notes_col_width = 44
    chars_per_line = max(10, int(item_col_width * 1.1))
    chars_notes = max(10, int(notes_col_width * 1.1))
    base_line_height = 15

    def _wrapped_line_count(text: str, cpl: int) -> int:
        raw_lines = text.splitlines() if text else [""]
        n = 0
        for ln in raw_lines:
            ln_len = len(ln)
            n += max(1, (ln_len + cpl - 1) // cpl)
        return n

    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 24
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            header = ws.cell(1, c).value
            if header in wrap_cols:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top")
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="FAFAFA")
            cell.border = thin_border

        line_count = 0
        if item_col_idx:
            v = ws.cell(r, item_col_idx).value
            text = "" if v is None else str(v)
            line_count = max(line_count, _wrapped_line_count(text, chars_per_line))
        if notes_col_idx:
            v2 = ws.cell(r, notes_col_idx).value
            t2 = "" if v2 is None else str(v2)
            line_count = max(line_count, _wrapped_line_count(t2, chars_notes))
        if line_count:
            ws.row_dimensions[r].height = max(24, min(300, base_line_height * line_count + 12))


def _format_pick_summary_sheet(ws: object) -> None:
    """출고_피킹요약 시트 가독성."""
    header_fill = PatternFill("solid", fgColor="E3F2FD")
    header_font = Font(bold=True)
    thin_side = Side(style="thin", color="B0B0B0")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_side = Side(style="medium", color="607D8B")
    header_border = Border(left=header_side, right=header_side, top=header_side, bottom=header_side)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 72
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = max(18, min(120, 14 * (1 + str(ws.cell(r, 2).value or "").count("\n"))))


def _picking_notes_excel_cell(o: pd.Series) -> str:
    """피킹 시트 '특이사항' 열: 아이콘 줄 + 엑셀 자동(attention_note) + 수기(special_issue)만.

    `order_list`는 같은 시트 **품목** 열에만 두고, `attention_note`/`special_issue` 안에 붙어 있는 동일 문구는 제거한다.
    """
    parts: list[str] = []
    ol = o.get("order_list")
    att_blk = _attention_note_export_for_order(o.get("attention_note"), ol)
    if att_blk:
        parts.append(att_blk)
    si = _strip_order_list_overlap(o.get("special_issue"), ol).strip()
    if si:
        parts.append(si)
    return "\n\n".join(parts) if parts else ""


def _build_shipped_excel_bytes(orders_all: pd.DataFrame, items_all: pd.DataFrame, shipped_date: dt.date) -> bytes:
    shipped = orders_all.copy()
    if "status" in shipped.columns:
        shipped = shipped[shipped["status"] == "출고"]
    # Filter to "today shipped" when shipped_at exists
    if "shipped_at" in shipped.columns:
        shipped["_shipped_date"] = _shipped_at_calendar_date_series(shipped["shipped_at"])
        shipped = shipped[shipped["_shipped_date"] == shipped_date].drop(columns=["_shipped_date"], errors="ignore")
    shipped = shipped.reset_index(drop=True)

    # Picking list (print-friendly)
    def _item_line_from_view(view: pd.DataFrame) -> str:
        lines: list[str] = []

        def _sv(v: object) -> str:
            if v is None:
                return ""
            try:
                if pd.isna(v):
                    return ""
            except Exception:
                pass
            return str(v).strip().replace("\n", " ")

        for i, (_, r) in enumerate(view.iterrows(), start=1):
            name = _sv(r.get("품목")) or "(품목명 없음)"
            parts = []
            spec = _sv(r.get("규격/옵션"))
            size = _sv(r.get("사이즈"))
            shelf = _sv(r.get("책장색상"))
            leg = _sv(r.get("다리색상"))
            qty = _sv(r.get("개수"))
            if spec:
                parts.append(spec)
            if size:
                parts.append(size)
            if shelf:
                parts.append(f"책장:{shelf}")
            if leg:
                parts.append(f"다리:{leg}")
            if qty:
                parts.append(f"개수:{qty}")
            ship = _sv(r.get("배송"))
            if ship:
                parts.append(f"배송:{ship}")
            line = f"{i}. {name}"
            if parts:
                line += " — " + " / ".join(parts)
            lines.append(line)
        return "\n".join(lines)

    def _picking_rows_for_kind(pick_kind: str) -> pd.DataFrame:
        rows: list[dict[str, object]] = []
        n = 0
        for _, o in shipped.iterrows():
            oid = o.get("order_id")
            view = _items_view_excel_pick_kind(items_all, str(oid), pick_kind) if oid is not None else pd.DataFrame()
            if len(view) == 0:
                continue
            n += 1
            rows.append(
                {
                    "No": n,
                    "받는분": o.get("receiver_name"),
                    "배송(엑셀)": _excel_ship_for_picking_row(oid, items_all, pick_kind),
                    "전화": o.get("phone"),
                    "주소": o.get("address"),
                    "배송요청": o.get("delivery_request"),
                    "특이사항": _picking_notes_excel_cell(o),
                    "품목": _item_line_from_view(view),
                }
            )
        cols = ["No", "받는분", "배송(엑셀)", "전화", "주소", "배송요청", "특이사항", "품목"]
        return pd.DataFrame(rows, columns=cols) if rows else pd.DataFrame(columns=cols)

    picking_taek = _picking_rows_for_kind("택배")
    picking_direct = _picking_rows_for_kind("직접배송")

    n_u, n_t_rows, n_d_rows, n_m = _picking_stats(shipped, items_all)
    mixed_labels: list[str] = []
    for _, o in shipped.iterrows():
        ht, hd = _order_picking_sheet_hits(items_all, o.get("order_id"))
        if ht and hd:
            rn = str(o.get("receiver_name") or "").strip()
            mixed_labels.append(rn or str(o.get("order_id") or ""))
    mixed_text = ", ".join(mixed_labels) if mixed_labels else "—"

    summary_df = pd.DataFrame(
        [
            {"항목": "출고 기준일", "값": shipped_date.isoformat()},
            {"항목": "출고 주문 수(주문그룹)", "값": n_u},
            {"항목": "피킹리스트_택배 행 수", "값": n_t_rows},
            {"항목": "피킹리스트_직접 행 수", "값": n_d_rows},
            {"항목": "혼합 배송 주문 수", "값": n_m},
            {
                "항목": "검산",
                "값": f"택배 {n_t_rows}행 + 직접 {n_d_rows}행 = 출고 {n_u}건 + 혼합 {n_m}건",
            },
            {"항목": "혼합 주문(받는분)", "값": mixed_text},
            {
                "항목": "안내",
                "값": "품목 배송란에 직접·택배가 함께 있는 주문은 두 피킹 시트에 각 한 줄씩 들어갑니다. 행 수 합은 출고 건수와 다를 수 있습니다.",
            },
        ]
    )

    # 로젠택배 서식: 택배로 나가는 주문만 (직접배송 전용 주문 제외)
    def _lozen_item_name_taek(o: pd.Series) -> str:
        oid = o.get("order_id")
        view = _items_view_excel_pick_kind(items_all, str(oid), "택배") if oid is not None else pd.DataFrame()
        if len(view) and "품목" in view.columns:
            first = str(view.iloc[0]["품목"]).strip()
            if first:
                return first
        ol = str(o.get("order_list") or "").strip().replace("\n", " / ")
        return (ol[:50] + "…") if len(ol) > 51 else ol

    lozen_rows: list[dict[str, object]] = []
    for _, o in shipped.iterrows():
        oid = o.get("order_id")
        ht, _hd = _order_picking_sheet_hits(items_all, oid)
        if not ht:
            continue
        lozen_rows.append(
            {
                "수하인명": o.get("receiver_name"),
                "수하인주소": o.get("address"),
                "수하인전화번호": o.get("phone"),
                "수하인휴대폰번호": o.get("phone"),
                "박스수량": 1,
                "택배운임": 3000,
                "운임구분": "선불",
                "품목명": _lozen_item_name_taek(o),
                "배송메세지": o.get("delivery_request"),
            }
        )
    lozen_cols = [
        "수하인명",
        "수하인주소",
        "수하인전화번호",
        "수하인휴대폰번호",
        "박스수량",
        "택배운임",
        "운임구분",
        "품목명",
        "배송메세지",
    ]
    lozen_sheet = pd.DataFrame(lozen_rows, columns=lozen_cols) if lozen_rows else pd.DataFrame(columns=lozen_cols)

    if "order_id" in shipped.columns and len(items_all) and "order_id" in items_all.columns:
        _sm = infer_settlement_ship_series(items_all)
        shipped = shipped.copy()
        shipped["배송(엑셀)"] = shipped["order_id"].astype(str).map(_sm).fillna("택배")
        shipped["배송(엑셀)"] = shipped["배송(엑셀)"].map(_ship_display_label)

    # Order sheet (one row per order)
    order_cols = [
        "purchase_date",
        "receiver_name",
        "phone",
        "address",
        "delivery_request",
        "attention_note",
        "order_list",
        "special_issue",
        "배송(엑셀)",
        "status",
        "shipped_at",
        "order_id",
        "source_file",
        "group_no",
        "group_start_row",
        "deadline_raw",
        "order_date_raw",
        "created_at",
    ]
    order_sheet = shipped[[c for c in order_cols if c in shipped.columns]].copy()
    if "attention_note" in order_sheet.columns:
        if "order_list" in order_sheet.columns:

            def _order_sheet_att_row(r: pd.Series) -> str:
                return _attention_note_export_for_order(r.get("attention_note"), r.get("order_list"))

            order_sheet["attention_note"] = order_sheet.apply(_order_sheet_att_row, axis=1)
        else:
            order_sheet["attention_note"] = order_sheet["attention_note"].map(_attention_note_with_icon_export)

    if "special_issue" in order_sheet.columns and "order_list" in order_sheet.columns:
        order_sheet["special_issue"] = order_sheet.apply(
            lambda r: _strip_order_list_overlap(r.get("special_issue"), r.get("order_list")).strip(),
            axis=1,
        )

    if "shipped_at" in order_sheet.columns:

        def _fmt_shipped_at_kst(v: object) -> str:
            if v is None:
                return ""
            if isinstance(v, str) and not v.strip():
                return ""
            try:
                if isinstance(v, float) and pd.isna(v):
                    return ""
            except Exception:
                pass
            t = pd.to_datetime(v, errors="coerce")
            if pd.isna(t):
                return str(v).strip()
            ts = pd.Timestamp(t)
            try:
                if ts.tzinfo is None:
                    ts2 = ts.tz_localize(_SEOUL, ambiguous="NaT", nonexistent="NaT")
                else:
                    ts2 = ts.tz_convert(_SEOUL)
            except Exception:
                ts2 = ts
            if pd.isna(ts2):
                return str(v).strip()
            return pd.Timestamp(ts2).strftime("%Y-%m-%d %H:%M:%S") + " (KST)"

        order_sheet["shipped_at"] = order_sheet["shipped_at"].map(_fmt_shipped_at_kst)

    # Items sheet: 준비할 품목/수량만 (품목별 합산)
    if len(shipped) and "order_id" in shipped.columns and "order_id" in items_all.columns:
        item_sheet = items_all[items_all["order_id"].isin(shipped["order_id"])].copy()
    else:
        item_sheet = items_all.head(0).copy()

    def _sv(v: object) -> str:
        if v is None:
            return ""
        try:
            if pd.isna(v):
                return ""
        except Exception:
            pass
        return str(v).strip().replace("\n", " ")

    if len(item_sheet):
        def _item_key(r: pd.Series) -> str:
            name = _sv(r.get("product_canonical")) or _sv(r.get("product_raw")) or "(품목명 없음)"
            parts = []
            spec = _sv(r.get("spec_raw"))
            note = _sv(r.get("note_raw"))
            size = _sv(r.get("size"))
            shelf = _sv(r.get("shelf_color"))
            leg = _sv(r.get("leg_color"))
            if spec:
                parts.append(spec)
            if note and note not in spec:
                parts.append(note)
            if size:
                parts.append(size)
            if shelf:
                parts.append(f"책장:{shelf}")
            if leg:
                parts.append(f"다리:{leg}")
            return f"{name} — " + " / ".join([p for p in parts if p]) if parts else name

        item_sheet["_품목"] = item_sheet.apply(_item_key, axis=1)
        item_sheet["_수량"] = pd.to_numeric(item_sheet.get("qty"), errors="coerce").fillna(0).astype(int)

        def _ship_join_for_export(s: pd.Series) -> str:
            """같은 품목 키로 묶인 행들의 배송란 → 출고_품목에는 항상 '직접'·'택배'만 표기(미기재·판별불가는 택배)."""
            kinds: set[str] = set()
            for x in s.tolist():
                if x is None:
                    continue
                try:
                    if isinstance(x, float) and pd.isna(x):
                        continue
                except Exception:
                    pass
                t = str(x).strip()
                if not t:
                    continue
                c = classify_ship_raw(x)
                if c == "직접배송":
                    kinds.add("직접")
                elif c == "택배":
                    kinds.add("택배")
                else:
                    kinds.add("택배")
            if not kinds:
                return "택배"
            order = ("직접", "택배")
            return " / ".join(k for k in order if k in kinds)

        item_sheet = (
            item_sheet.groupby("_품목", as_index=False)
            .agg(_수량=("_수량", "sum"), 엑셀배송=("ship_raw", _ship_join_for_export))
            .rename(columns={"_품목": "품목", "_수량": "수량"})
            .sort_values(["품목"])
            .reset_index(drop=True)
        )
        item_sheet["엑셀배송"] = item_sheet["엑셀배송"].replace("", "택배").fillna("택배")
    else:
        item_sheet = pd.DataFrame(columns=["품목", "수량", "엑셀배송"])

    # Friendly column names
    order_sheet = order_sheet.rename(
        columns={
            "purchase_date": "구매일자",
            "receiver_name": "받는분",
            "phone": "전화",
            "address": "주소",
            "delivery_request": "배송요청",
            "attention_note": "특이사항(엑셀자동)",
            "order_list": "주문목록",
            "special_issue": "특이사항",
            "status": "상태",
            "shipped_at": "출고처리시각",
            "order_id": "주문ID",
            "source_file": "원본파일",
            "group_no": "번호",
            "group_start_row": "그룹시작행",
            "deadline_raw": "납기",
            "order_date_raw": "엑셀주문일자",
            "created_at": "수집시각",
        }
    )
    # item_sheet columns are already Korean: 품목, 수량

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="출고_피킹요약")
        picking_taek.to_excel(writer, index=False, sheet_name="피킹리스트_택배")
        picking_direct.to_excel(writer, index=False, sheet_name="피킹리스트_직접")
        lozen_sheet.to_excel(writer, index=False, sheet_name="로젠택배")
        order_sheet.to_excel(writer, index=False, sheet_name="출고_주문")
        item_sheet.to_excel(writer, index=False, sheet_name="출고_품목")

        header_fill = PatternFill("solid", fgColor="E3F2FD")
        header_font = Font(bold=True)
        thin_side = Side(style="thin", color="B0B0B0")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        header_side = Side(style="medium", color="607D8B")
        header_border = Border(left=header_side, right=header_side, top=header_side, bottom=header_side)

        _format_pick_summary_sheet(writer.book["출고_피킹요약"])
        _format_picking_worksheet(writer.book["피킹리스트_택배"])
        _format_picking_worksheet(writer.book["피킹리스트_직접"])

        # Formatting for 로젠택배 sheet (A4 landscape + grid)
        ws_l = writer.book["로젠택배"]
        ws_l.page_setup.orientation = "landscape"
        ws_l.page_setup.paperSize = ws_l.PAPERSIZE_A4
        ws_l.page_setup.fitToWidth = 1
        ws_l.page_setup.fitToHeight = 0
        ws_l.print_title_rows = "1:1"
        ws_l.page_margins.left = 0.25
        ws_l.page_margins.right = 0.25
        ws_l.page_margins.top = 0.35
        ws_l.page_margins.bottom = 0.35
        ws_l.sheet_properties.pageSetUpPr.fitToPage = True
        ws_l.freeze_panes = "A2"
        ws_l.auto_filter.ref = ws_l.dimensions
        for cell in ws_l[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = header_border

        # Column widths close to Lozen template
        widths = {
            "수하인명": 12,
            "수하인주소": 48,
            "수하인전화번호": 14,
            "수하인휴대폰번호": 14,
            "박스수량": 8,
            "택배운임": 10,
            "운임구분": 8,
            "품목명": 20,
            "배송메세지": 40,
        }
        header_map_l = {ws_l.cell(1, c).value: c for c in range(1, ws_l.max_column + 1)}
        for h, col_idx in header_map_l.items():
            letter = ws_l.cell(1, col_idx).column_letter
            if h in widths:
                ws_l.column_dimensions[letter].width = widths[h]

        wrap_l = {"수하인주소", "배송메세지"}
        for r in range(2, ws_l.max_row + 1):
            ws_l.row_dimensions[r].height = 36
            for c in range(1, ws_l.max_column + 1):
                cell = ws_l.cell(r, c)
                header = ws_l.cell(1, c).value
                if header in wrap_l:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
                cell.border = thin_border
                if r % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="FAFAFA")

        # Formatting for item sheet (prepare list)
        ws2 = writer.book["출고_품목"]
        # Print setup: A4 landscape, fit to page width
        ws2.page_setup.orientation = "landscape"
        ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
        ws2.page_setup.fitToWidth = 1
        ws2.page_setup.fitToHeight = 0
        ws2.print_title_rows = "1:1"
        ws2.page_margins.left = 0.25
        ws2.page_margins.right = 0.25
        ws2.page_margins.top = 0.35
        ws2.page_margins.bottom = 0.35
        ws2.page_margins.header = 0.2
        ws2.page_margins.footer = 0.2
        ws2.sheet_properties.pageSetUpPr.fitToPage = True

        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = ws2.dimensions
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = header_border

        # Column widths + wrap for long text
        wrap_cols2 = {"품목", "엑셀배송"}
        header_map2 = {ws2.cell(1, c).value: c for c in range(1, ws2.max_column + 1)}
        for name, col_idx in header_map2.items():
            letter = ws2.cell(1, col_idx).column_letter
            if name == "품목":
                ws2.column_dimensions[letter].width = 70
            elif name == "수량":
                ws2.column_dimensions[letter].width = 8
            elif name == "엑셀배송":
                ws2.column_dimensions[letter].width = 18

        for r in range(2, ws2.max_row + 1):
            ws2.row_dimensions[r].height = 28
            for c in range(1, ws2.max_column + 1):
                cell = ws2.cell(r, c)
                header = ws2.cell(1, c).value
                if header in wrap_cols2:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
                if r % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="FAFAFA")
                cell.border = thin_border
    return buf.getvalue()


def _build_lozen_xlsx_bytes_for_orders(
    orders_all: pd.DataFrame,
    items_all: pd.DataFrame,
    order_ids: list[str],
) -> bytes:
    """선택된 주문들로 '로젠택배' 업로드용 단일 시트 엑셀을 만든다."""
    cols = [
        "수하인명",
        "수하인주소",
        "수하인전화번호",
        "수하인휴대폰번호",
        "박스수량",
        "택배운임",
        "운임구분",
        "품목명",
        "배송메세지",
    ]

    if orders_all is None or len(orders_all) == 0 or not order_ids:
        buf0 = io.BytesIO()
        with pd.ExcelWriter(buf0, engine="openpyxl") as w:
            pd.DataFrame(columns=cols).to_excel(w, index=False, sheet_name="로젠택배")
        return buf0.getvalue()

    sub = orders_all[orders_all["order_id"].astype(str).isin([str(x) for x in order_ids])].copy().reset_index(drop=True)

    # 택배만 포함 (엑셀 배송란 추정)
    if items_all is not None and len(items_all) and "order_id" in items_all.columns:
        ship_series = infer_settlement_ship_series(items_all)
        ship_kind = sub["order_id"].astype(str).map(ship_series).fillna("택배")
        sub = sub[ship_kind.astype(str).isin(["택배"])].copy()

    def _item_name(o: pd.Series) -> str:
        ol = str(o.get("order_list") or "").strip().replace("\n", " / ")
        if not ol:
            return ""
        return (ol[:50] + "…") if len(ol) > 51 else ol

    lozen = pd.DataFrame(
        [
            {
                "수하인명": o.get("receiver_name"),
                "수하인주소": o.get("address"),
                "수하인전화번호": o.get("phone"),
                "수하인휴대폰번호": o.get("phone"),
                "박스수량": 1,
                "택배운임": 3000,
                "운임구분": "선불",
                "품목명": _item_name(o),
                "배송메세지": o.get("delivery_request"),
            }
            for _, o in sub.iterrows()
        ],
        columns=cols,
    )

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        lozen.to_excel(w, index=False, sheet_name="로젠택배")
    return buf.getvalue()


def update_special_issue_for_orders(db_path: str, order_ids: list[str], text: str) -> None:
    if not order_ids:
        return
    con = sqlite3.connect(db_path)
    try:
        cur = con.cursor()
        cur.executemany(
            "UPDATE orders SET special_issue=? WHERE order_id=?",
            [(text, oid) for oid in order_ids],
        )
        con.commit()
    finally:
        con.close()


def init_shared_session_state() -> None:
    if "mutomo_db_reload_nonce" not in st.session_state:
        st.session_state["mutomo_db_reload_nonce"] = 0
    if "active_selector" not in st.session_state:
        st.session_state["active_selector"] = None  # "search" | "table"
    if "search_pick_ids" not in st.session_state:
        st.session_state["search_pick_ids"] = []
    if "selected_ids_from_table" not in st.session_state:
        st.session_state["selected_ids_from_table"] = []
    if "prev_search_pick_ids" not in st.session_state:
        st.session_state["prev_search_pick_ids"] = []
    if "prev_table_pick_ids" not in st.session_state:
        st.session_state["prev_table_pick_ids"] = []
    if "request_clear_search_pick_labels" not in st.session_state:
        st.session_state["request_clear_search_pick_labels"] = False
    if "table_sync_from_search" not in st.session_state:
        st.session_state["table_sync_from_search"] = False


# 접수로 되돌리기 전 확인(메인 영역에 표시 — 사이드바 맨 아래에 묻히지 않게)
MUTOMO_PENDING_REVERT_KEY = "mutomo_pending_revert_received"

# 사이드바 상태 필터·버튼에서 공통 사용 (SQLite `orders.status` 텍스트)
ORDER_STATUS_FILTER_OPTIONS: tuple[str, ...] = (
    "접수",
    "도면참조",
    "클레임",
    "출고",
    "마감",
    "납품취소",
)
# 접수로 되돌릴 때는 '도면만 받은 단계(도면참조)'까지는 경고 없이(접수와 동일 취급)
_PRE_SHIP_STATUSES_FOR_REVERT: frozenset[str] = frozenset({"접수", "도면참조"})


def _revert_to_received_warn_message(sub: pd.DataFrame) -> tuple[bool, str]:
    """접수로 되돌리기 전 경고가 필요한지, 요약 문구(한 줄)를 돌려준다."""
    if sub is None or len(sub) == 0:
        return False, ""
    parts: list[str] = []
    need = False
    if "status" in sub.columns:
        stt = sub["status"].astype(str).str.strip()
        other = stt[~stt.isin(_PRE_SHIP_STATUSES_FOR_REVERT)]
        if len(other):
            need = True
            vc = other.value_counts()
            for k, v in vc.items():
                parts.append(f"{k} {int(v)}건")
    ts_parts: list[str] = []
    if "shipped_at" in sub.columns:
        sa = sub["shipped_at"].astype(str).fillna("").str.strip()
        n = int(((sa != "") & (sa.str.lower() != "nan")).sum())
        if n:
            need = True
            ts_parts.append(f"출고시각 있음 {n}건")
    if "closed_at" in sub.columns:
        ca = sub["closed_at"].astype(str).fillna("").str.strip()
        n = int(((ca != "") & (ca.str.lower() != "nan")).sum())
        if n:
            need = True
            ts_parts.append(f"마감시각 있음 {n}건")
    msg = " / ".join([*parts, *ts_parts])
    return need, msg


# 이름 검색 → 전체 접수 표 선택 연동 시 한 번에 넘길 최대 행 수 (브라우저·Streamlit 부담 완화)
_MAX_SEARCH_TO_TABLE_ROWS = 2000

# 설정 사이드바「최근 수집 파일」표: DB에 있는 소스 파일 전체가 아니라, 마지막 수집 시각 기준 상위 N건만 미리보기
_RECENT_SOURCE_FILES_SIDEBAR_PREVIEW = 50


def _unique_receiver_names(orders: pd.DataFrame) -> list[str]:
    """비어 있지 않은 받는분 이름을 대략 등장 순으로 유일하게 모은다."""
    if orders.empty or "receiver_name" not in orders.columns:
        return []
    s = orders["receiver_name"].astype(str).str.strip()
    s = s[s != ""]
    seen: set[str] = set()
    out: list[str] = []
    for x in s:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def _focus_streamlit_input_by_placeholder(placeholder: str) -> None:
    """Streamlit에 네이티브 포커스 API가 없어, '다음 이름' 등 placeholder로 입력칸을 찾아 포커스."""
    import html

    ph = html.escape(placeholder, quote=True)
    st_components.html(
        f"""<script>
(function () {{
  const root = window.parent.document;
  const el = root.querySelector('input[placeholder="{ph}"]');
  if (el) {{
    el.focus();
    el.scrollIntoView({{ block: "nearest", behavior: "instant" }});
  }}
}})();
</script>""",
        height=1,
        width=1,
    )


def _fuzzy_receiver_name_suggestions(part: str, candidates: list[str], *, limit: int = 10) -> list[str]:
    """부분 문자열 검색이 0건일 때 오타 완화용 제안(한글·영문 모두 difflib 기준)."""
    q = part.strip()
    if len(q) < 2 or not candidates:
        return []
    # 1차·2차: difflib 근접 매칭 (한글 오타는 cutoff를 한 단계 낮춰도 시도)
    hits = difflib.get_close_matches(q, candidates, n=limit, cutoff=0.48)
    if not hits:
        hits = difflib.get_close_matches(q, candidates, n=limit, cutoff=0.40)
    if hits:
        return hits
    # 2차: 유일 이름 수가 적을 때만 비율 상위 (전수는 수만 건에서 부담)
    cap = 2500
    if len(candidates) > cap:
        return []
    scored: list[tuple[float, str]] = []
    for c in candidates:
        r = difflib.SequenceMatcher(a=q, b=c, autojunk=False).ratio()
        if r >= 0.38:
            scored.append((r, c))
    scored.sort(key=lambda t: (-t[0], t[1]))
    return [c for _r, c in scored[:limit]]


def _receiver_name_match_count(orders: pd.DataFrame, term: str) -> int:
    """받는분 이름에 term(부분일치)이 포함된 행 수."""
    t = (term or "").strip()
    if not t or orders.empty or "receiver_name" not in orders.columns:
        return 0
    ns = orders["receiver_name"].astype(str)
    return int(ns.str.contains(t, case=False, na=False, regex=False).sum())


def _receiver_name_or_match_count(orders: pd.DataFrame, parts_list: list[str]) -> int:
    """parts_list 전체를 OR(부분일치)로 합친 접수 건수."""
    if not parts_list or orders.empty or "receiver_name" not in orders.columns:
        return 0
    nm = orders["receiver_name"].astype(str)
    mask = nm.str.contains(parts_list[0], case=False, na=False, regex=False)
    for t in parts_list[1:]:
        mask = mask | nm.str.contains(t, case=False, na=False, regex=False)
    return int(mask.sum())


def _identity_hint_lines(hits: pd.DataFrame) -> list[str]:
    """검색 결과에 동명이인·재주문 패턴이 있으면 안내 문장 목록."""
    lines: list[str] = []
    if hits.empty or "receiver_name" not in hits.columns:
        return lines
    hn = hits.copy()
    hn["_rn"] = hn["receiver_name"].astype(str).str.strip()
    hn = hn[hn["_rn"] != ""]
    if hn.empty:
        return lines

    if "party_key" in hn.columns:
        pk = hn["party_key"].fillna("").astype(str)
        for name, grp in hn.groupby("_rn", sort=False):
            if len(grp) < 2:
                continue
            u = {x for x in pk.loc[grp.index].tolist() if x and x != "nan"}
            if len(u) >= 2:
                lines.append(
                    f"「{name}」은(는) **동명이인**일 가능성이 큽니다. (연락처·주소 기준으로 **{len(u)}무리**로 갈립니다.)"
                )
        seen_pk: set[str] = set()
        for pkey, grp in hn.groupby(pk, sort=False):
            ps = str(pkey).strip()
            if len(grp) < 2 or not ps or ps.startswith("order:"):
                continue
            if ps in seen_pk:
                continue
            seen_pk.add(ps)
            nm = str(grp.iloc[0].get("receiver_name") or "").strip() or "동일 연락처"
            lines.append(
                f"「{nm}」은(는) **재주문**으로 보입니다. (**동일 연락처** 접수가 **{len(grp)}건** 묶여 있습니다.)"
            )
    else:
        vc = hn["_rn"].value_counts()
        for name, c in vc.items():
            if int(c) >= 2:
                lines.append(
                    f"「{name}」이(가) **{int(c)}건** 있습니다. **동명이인**·**재주문**·여러 접수일 수 있으니 전화·날짜로 구분해 주세요."
                )

    out: list[str] = []
    for x in lines:
        if x not in out:
            out.append(x)
    return out[:12]


def _identity_hint_tags(hits: pd.DataFrame) -> list[str]:
    """검색 결과(한 검색어 범위)에 동명이인·재주문 패턴이 있으면 짧은 태그 목록."""
    tags: list[str] = []
    if hits.empty or "receiver_name" not in hits.columns:
        return tags
    hn = hits.copy()
    hn["_rn"] = hn["receiver_name"].astype(str).str.strip()
    hn = hn[hn["_rn"] != ""]
    if hn.empty:
        return tags

    if "party_key" in hn.columns:
        pk = hn["party_key"].fillna("").astype(str)
        for _name, grp in hn.groupby("_rn", sort=False):
            if len(grp) < 2:
                continue
            u = {x for x in pk.loc[grp.index].tolist() if x and x != "nan"}
            if len(u) >= 2 and "동명이인" not in tags:
                tags.append("동명이인")
        seen_pk: set[str] = set()
        for pkey, grp in hn.groupby(pk, sort=False):
            ps = str(pkey).strip()
            if len(grp) < 2 or not ps or ps.startswith("order:"):
                continue
            if ps in seen_pk:
                continue
            seen_pk.add(ps)
            if "재주문" not in tags:
                tags.append("재주문")
    else:
        vc = hn["_rn"].value_counts()
        if (vc >= 2).any() and "동명이인" not in tags:
            tags.append("동명이인")

    return tags


def _clear_row_pick_label_keys() -> None:
    for k in list(st.session_state.keys()):
        if isinstance(k, str) and k.startswith("row_pick_labels_"):
            st.session_state.pop(k, None)


def _clear_name_search_row_input_keys() -> None:
    for k in list(st.session_state.keys()):
        if isinstance(k, str) and k.startswith("name_search_row_input_"):
            st.session_state.pop(k, None)


def _prune_name_search_row_input_keys(n_terms: int) -> None:
    for k in list(st.session_state.keys()):
        if not isinstance(k, str) or not k.startswith("name_search_row_input_"):
            continue
        suf = k.removeprefix("name_search_row_input_")
        try:
            j = int(suf)
        except ValueError:
            st.session_state.pop(k, None)
            continue
        if j >= n_terms:
            st.session_state.pop(k, None)


def _build_order_picker_lists(
    rows: pd.DataFrame, items_df: pd.DataFrame | None = None
) -> tuple[dict[str, str], list[str], list[str]]:
    """행별 multiselect용 (라벨→order_id, 옵션, 기본 선택). rows는 정렬된 검색 결과."""
    if rows.empty or "order_id" not in rows.columns:
        return {}, [], []
    r = rows.copy()
    rkey = r["receiver_name"].astype(str).str.strip() if "receiver_name" in r.columns else pd.Series([""] * len(r))
    _vc = rkey.value_counts()
    r["_homonym_hit"] = rkey.map(lambda x: int(_vc.get(x, 0) > 1))

    def _digits_tail(phone_val: object, n: int = 4) -> str:
        s = "".join(ch for ch in str(phone_val or "") if ch.isdigit())
        return s[-n:] if len(s) >= n else ""

    def _label(row: pd.Series) -> str:
        purchase = str(row.get("purchase_date") or "").strip()
        name = str(row.get("receiver_name") or "").strip()
        status = str(row.get("status") or "").strip()
        att_stripped = _strip_order_list_overlap(row.get("attention_note"), row.get("order_list"))
        att_src = _combined_attention_for_icons(att_stripped or None, row.get("special_issue"))
        name_seg = _compact_name_display(
            status, name, att_src, sheet_blob=_drawing_ref_sheet_blob(row, items_df)
        )
        order_list = str(row.get("order_list") or "").strip().replace("\n", " / ")
        order_list = (order_list[:120] + "…") if len(order_list) > 121 else order_list
        base = f"{purchase} | {name_seg} | {order_list}".strip(" |")
        if int(row.get("_homonym_hit") or 0):
            tail = _digits_tail(row.get("phone"), 4)
            if tail:
                base = f"{base} ·전화끝{tail}".strip()
            else:
                oid = str(row.get("order_id") or "").strip()
                if oid:
                    base = f"{base} ·ID {oid[:10]}".strip()
        return base

    r["_label"] = r.apply(_label, axis=1)
    dup = r["_label"].duplicated(keep=False)
    if dup.any():
        r.loc[dup, "_label"] = r.loc[dup].apply(lambda row: f"{row['_label']} ({row['order_id']})", axis=1)

    label_to_id = dict(zip(r["_label"].tolist(), r["order_id"].tolist(), strict=False))
    options = r["_label"].tolist()
    default_lbl: list[str] = []
    if "status" in r.columns:
        for lbl in options:
            oid = str(label_to_id.get(lbl, ""))
            hit = r[r["order_id"].astype(str) == oid]
            if len(hit) and str(hit.iloc[0].get("status") or "").strip() != "출고":
                default_lbl = [lbl]
                break
    if not default_lbl and options:
        default_lbl = [options[0]]
    return label_to_id, options, default_lbl


def render_receiver_name_search(
    orders: pd.DataFrame,
    items: pd.DataFrame,
    db_path: str,
    *,
    orders_name_hints: pd.DataFrame | None = None,
) -> None:
    """이름(좁은 칸) + 옆 multiselect(라벨 숨김, 줄 맞춤). 다음 줄은 폼에서 Enter로 추가. OR 합쳐 출고."""
    # 품목줄까지 포함한 도면참조 표시(📐)에 사용. 상세 UI는 메인에서 제거됨.
    nh = orders_name_hints if orders_name_hints is not None else orders
    st.subheader("검색")
    if "name_search_terms" not in st.session_state:
        st.session_state["name_search_terms"] = []
    # text_input(key=...) 생성 이후에는 같은 런에서 해당 키를 쓸 수 없음 → 이전 런에서 요청한 초기화만 여기서 처리
    if st.session_state.pop("_reset_name_search_draft", False):
        st.session_state.pop("name_search_draft", None)
        st.session_state.pop("name_search_draft_next", None)
        st.session_state.pop("name_search_inline_next_draft", None)
        for _k in list(st.session_state.keys()):
            if isinstance(_k, str) and _k.startswith("name_inline_draft_"):
                st.session_state.pop(_k, None)
        _clear_name_search_row_input_keys()

    terms: list[str] = st.session_state["name_search_terms"]
    parts = [str(p).strip() for p in terms if str(p).strip()]

    nm_series = orders.get("receiver_name", "").astype(str)
    if parts:
        mask = nm_series.str.contains(parts[0], case=False, na=False, regex=False)
        for q in parts[1:]:
            mask = mask | nm_series.str.contains(q, case=False, na=False, regex=False)
        hits = orders[mask].copy()
    else:
        hits = orders.iloc[:0].copy()

    if not parts:
        with st.form("name_search_first_row", clear_on_submit=True):
            c1, c2 = st.columns([1, 7], gap="small")
            with c1:
                st.text_input(
                    "이름",
                    key="name_search_draft",
                    placeholder="이름",
                    label_visibility="collapsed",
                )
            with c2:
                submitted_first = st.form_submit_button("⏎", type="secondary", use_container_width=False)
        if submitted_first:
            q = (st.session_state.get("name_search_draft") or "").strip()
            if q and q not in st.session_state["name_search_terms"]:
                st.session_state["name_search_terms"].append(q)
                if _receiver_name_match_count(orders, q) == 0:
                    st.session_state["_focus_next_name_inline"] = True
                st.rerun()

    search_block_on = bool(parts)
    pick_ids: list[str] = []
    if not search_block_on:
        st.divider()
        st.session_state["search_pick_ids"] = []
        st.session_state["prev_search_pick_ids"] = []
        st.session_state.pop("search_pick_labels", None)
        st.session_state.pop("_search_sync_order_ids", None)
        st.session_state.pop("_search_sync_sig", None)
        st.session_state["table_sync_from_search"] = False
        # 검색어가 없을 때도 아래 "전체 접수 목록" 표는 그대로 두어야 함.
        # 여기서 mutomo_full_orders_df 선택을 매 리런마다 비우면, 이름 검색 없이 표만 클릭해도 선택이 즉시 풀린다.
        return

    st.divider()
    hdr_l, hdr_m = st.columns([1, 7], gap="small")
    with hdr_l:
        st.markdown("")
    with hdr_m:
        b_pop, b_clr = st.columns(2, gap="small")
        with b_pop:
            pop_clicked = st.button(
                "맨 끝만 삭제",
                key="name_search_pop_btn",
                use_container_width=True,
                help="검색 조건에서 가장 마지막 줄만 삭제합니다.",
            )
        with b_clr:
            clear_clicked = st.button(
                "전부 비우기",
                key="name_search_clear_btn",
                use_container_width=True,
                help="검색 조건·입력·선택·아래 표 연동을 한꺼번에 초기화합니다.",
            )

    if pop_clicked:
        if st.session_state["name_search_terms"]:
            i_last = len(st.session_state["name_search_terms"]) - 1
            st.session_state["name_search_terms"].pop()
            st.session_state.pop(f"row_pick_labels_{i_last}", None)
            st.session_state.pop(f"name_search_row_input_{i_last}", None)
        if not st.session_state["name_search_terms"]:
            st.session_state["_search_ui_full_reset"] = True
        st.rerun()
    if clear_clicked:
        st.session_state["name_search_terms"] = []
        _clear_row_pick_label_keys()
        _clear_name_search_row_input_keys()
        for _k in list(st.session_state.keys()):
            if isinstance(_k, str) and _k.startswith("name_inline_draft_"):
                st.session_state.pop(_k, None)
        st.session_state["_reset_name_search_draft"] = True
        st.session_state["_search_ui_full_reset"] = True
        st.rerun()

    def _sort_hits_block(df: pd.DataFrame) -> pd.DataFrame:
        out = df.copy()
        if out.empty:
            return out
        if "status" in out.columns:
            out["_shipped_last"] = out["status"].astype(str).str.strip().eq("출고").astype(int)
            _sc = [c for c in ["_shipped_last", "purchase_date"] if c in out.columns]
            if _sc:
                out = out.sort_values(_sc, ascending=[True, False], na_position="last")
            out = out.drop(columns=["_shipped_last"], errors="ignore")
        elif "purchase_date" in out.columns:
            out = out.sort_values("purchase_date", ascending=False, na_position="last")
        return out

    if st.session_state.get("request_clear_search_pick_labels"):
        _clear_row_pick_label_keys()
        st.session_state.pop("search_pick_labels", None)
        st.session_state["request_clear_search_pick_labels"] = False

    terms_rows: list[str] = st.session_state["name_search_terms"]
    _prune_name_search_row_input_keys(len(terms_rows))

    # fuzzy 이름 버튼은 위젯 생성 이후라 같은 런에서 row_input 키를 못 바꿈 → 다음 런 초반에 반영
    _sug_apply = st.session_state.pop("_name_suggest_apply", None)
    if isinstance(_sug_apply, (list, tuple)) and len(_sug_apply) == 2:
        _sug_i, _sug_nm = _sug_apply
        if isinstance(_sug_i, int) and isinstance(_sug_nm, str):
            _sn2 = _sug_nm.strip()
            _ts = st.session_state["name_search_terms"]
            if _sn2 and 0 <= _sug_i < len(_ts):
                _ts[_sug_i] = _sn2
                st.session_state.pop(f"name_search_row_input_{_sug_i}", None)
                st.session_state.pop(f"row_pick_labels_{_sug_i}", None)

    def _row_term_live(idx: int) -> str:
        ts2 = st.session_state["name_search_terms"]
        if idx >= len(ts2):
            return ""
        rk = f"name_search_row_input_{idx}"
        base = str(ts2[idx]).strip()
        if rk not in st.session_state:
            return base
        return str(st.session_state[rk]).strip()

    def _row_term_for_match(idx: int) -> str:
        """위젯이 잠깐 비면 term_live가 ''가 되는데, pandas str.contains('')는 전 행 True라 라벨·선택이 깨짐 → 커밋된 줄 이름으로 보정."""
        ts2 = st.session_state["name_search_terms"]
        if idx >= len(ts2):
            return ""
        live = (_row_term_live(idx) or "").strip()
        if live:
            return live
        return str(ts2[idx]).strip()

    for i in range(len(terms_rows)):
        term0 = str(terms_rows[i]).strip()
        if not term0:
            continue
        rk = f"name_search_row_input_{i}"
        if rk not in st.session_state:
            st.session_state[rk] = term0
        term_live = _row_term_live(i)
        term_match = _row_term_for_match(i)
        row_hits = (
            orders.iloc[:0].copy()
            if not term_match
            else orders[nm_series.str.contains(term_match, case=False, na=False, regex=False)].copy()
        )
        row_hits = _sort_hits_block(row_hits)
        tags = _identity_hint_tags(row_hits)
        mk = f"row_pick_labels_{i}"
        _, options_r, default_lbl_r = _build_order_picker_lists(row_hits, items)
        if not options_r:
            # 0건일 때 pop 하면 다음 렌더에서 복귀해도 선택이 영구 사라짐 → 일시 0건은 키 유지
            if len(row_hits) > 0:
                st.session_state.pop(mk, None)
        else:
            if mk not in st.session_state:
                st.session_state[mk] = [] if tags else (list(default_lbl_r) if default_lbl_r else [])
            raw = st.session_state.get(mk, [])
            if not isinstance(raw, list):
                raw = []
            valid = [x for x in raw if x in options_r]
            if valid != raw:
                # 옵션 라벨이 한 프레임만 어긋나면 valid가 []가 되기 쉬움 → 기존 선택 통째로 비우지 않음
                if valid or not raw:
                    st.session_state[mk] = valid
            if not tags and not (st.session_state.get(mk) or []):
                st.session_state[mk] = list(default_lbl_r) if default_lbl_r else ([options_r[0]] if options_r else [])

        c_nm, c_ord = st.columns([1, 7], gap="small")
        with c_nm:
            st.text_input(
                "이름",
                key=rk,
                placeholder="이름",
                label_visibility="collapsed",
            )
        with c_ord:
            if options_r:
                st.multiselect(
                    " ",
                    options=options_r,
                    key=mk,
                    label_visibility="collapsed",
                )
            elif len(row_hits) == 0:
                if term_match and _receiver_name_match_count(orders, term_match) == 0 and _receiver_name_match_count(nh, term_match) > 0:
                    st.caption(
                        "전체 접수에는 이 이름이 있으나 **현재 사이드바 상태 필터**에는 없습니다. "
                        "필터에 **접수·출고** 등을 포함하면 오른쪽에서 고를 수 있습니다."
                    )
                _cand_row = _unique_receiver_names(nh)
                _sugs_row = _fuzzy_receiver_name_suggestions(term_match, _cand_row, limit=6)
                if term_match and _sugs_row:
                    st.caption("이 이름으로는 접수가 없습니다. **등록명**과 한 글자만 달라도 안 잡힙니다. 비슷한 이름:")
                    _nc = min(4, len(_sugs_row))
                    _rcols = st.columns(_nc)
                    for _si, _sn in enumerate(_sugs_row):
                        with _rcols[_si % _nc]:
                            if st.button(_sn, key=f"_row_name_suggest_{i}_{_si}"):
                                st.session_state["_name_suggest_apply"] = (i, _sn)
                                st.rerun()
                elif term_match:
                    st.caption("일치하는 접수가 없습니다.")
            else:
                st.caption("주문 목록을 만들 수 없습니다(데이터에 `order_id`가 있는지 확인).")

    def _row_pick_satisfied(idx: int) -> bool:
        t = _row_term_for_match(idx)
        if not t:
            return True
        rh = orders[nm_series.str.contains(t, case=False, na=False, regex=False)].copy()
        rh = _sort_hits_block(rh)
        mk2 = f"row_pick_labels_{idx}"
        picked = st.session_state.get(mk2, [])
        if len(rh) == 0:
            return True
        _, opts, _ = _build_order_picker_lists(rh, items)
        if not opts:
            return True
        return len(picked) >= 1

    can_add_next = all(_row_pick_satisfied(i) for i in range(len(terms_rows)))
    _rows_missing_pick: list[str] = []
    for _j in range(len(terms_rows)):
        if not _row_pick_satisfied(_j):
            _lab = (_row_term_for_match(_j) or str(terms_rows[_j]).strip() or f"{_j + 1}번째").strip()
            if _lab:
                _rows_missing_pick.append(_lab)

    if len(terms_rows) >= 1:
        _n_terms = len(terms_rows)
        _inline_draft_key = f"name_inline_draft_{_n_terms}"
        _inline_form_key = f"name_inline_form_{_n_terms}"
        st.session_state.pop("name_search_inline_next_draft", None)
        with st.form(_inline_form_key, clear_on_submit=True):
            a1, a2 = st.columns([1, 7], gap="small")
            with a1:
                st.text_input(
                    "이름",
                    key=_inline_draft_key,
                    placeholder="다음 이름",
                    label_visibility="collapsed",
                )
            with a2:
                submitted_inline = st.form_submit_button("⏎", type="secondary", use_container_width=False)
        if submitted_inline:
            qn = (st.session_state.get(_inline_draft_key) or "").strip()
            if not can_add_next:
                if qn:
                    if _rows_missing_pick:
                        _who = "**, **".join(_rows_missing_pick)
                        st.warning(
                            f"아래 이름 **검색 줄마다** 오른쪽에서 **처리할 주문**을 골라야 합니다. "
                            f"아직 선택이 없는 줄: **{_who}**."
                        )
                    else:
                        st.warning("검색 줄에서 **처리할 주문**을 먼저 고른 뒤에 다음 이름을 추가할 수 있습니다.")
            elif qn and qn not in st.session_state["name_search_terms"]:
                st.session_state["name_search_terms"].append(qn)
                if _receiver_name_match_count(orders, qn) == 0:
                    st.session_state["_focus_next_name_inline"] = True
                st.rerun()

    parts_live: list[str] = []
    for _i in range(len(terms_rows)):
        _t = _row_term_for_match(_i)
        if _t:
            parts_live.append(_t)
    if parts_live:
        mask_live = nm_series.str.contains(parts_live[0], case=False, na=False, regex=False)
        for _q in parts_live[1:]:
            mask_live = mask_live | nm_series.str.contains(_q, case=False, na=False, regex=False)
        hits_live = orders[mask_live].copy()
    else:
        hits_live = orders.iloc[:0].copy()

    st.markdown("**선택 요약**")
    pick_ids = []
    seen_oid: set[str] = set()
    for i in range(len(terms_rows)):
        tl = _row_term_for_match(i)
        if not tl:
            continue
        row_hits = orders[nm_series.str.contains(tl, case=False, na=False, regex=False)].copy()
        row_hits = _sort_hits_block(row_hits)
        ltd, _opts, _ = _build_order_picker_lists(row_hits, items)
        mk3 = f"row_pick_labels_{i}"
        for lbl in st.session_state.get(mk3, []):
            oid = str(ltd.get(lbl, "")).strip()
            if oid and oid not in seen_oid:
                seen_oid.add(oid)
                pick_ids.append(oid)

    n_sum = len(pick_ids)
    st.caption(f"**{n_sum}건**")

    _already_done_status = frozenset({"출고", "마감", "납품취소"})
    if pick_ids and "status" in orders.columns and "order_id" in orders.columns:
        _sp = orders[orders["order_id"].astype(str).isin([str(x) for x in pick_ids])]
        if len(_sp):
            _st = _sp["status"].astype(str).str.strip()
            _n_done = int(_st.isin(_already_done_status).sum())
            if _n_done:
                st.warning(
                    f"선택 **{n_sum}건** 중 **{_n_done}건**은 이미 처리된 주문입니다 "
                    f"(출고·마감·납품취소). **출고** 버튼은 출고 완료 건을 건너뜁니다."
                )

    if len(hits_live) == 0:
        pass

    _hint_lines = _identity_hint_lines(hits_live)
    if _hint_lines and len(hits_live):
        st.info("\n\n".join(_hint_lines))

    st.session_state["search_pick_ids"] = pick_ids

    if pick_ids != st.session_state.get("prev_search_pick_ids"):
        st.session_state["prev_search_pick_ids"] = pick_ids

    ship_action = st.button("선택 건 출고 처리", type="primary", key="ship_one_button")
    rc1, rc2, rc3, rc4, rc5 = st.columns([1, 1, 1, 1, 1.4])
    with rc1:
        claim_action = st.button("클레임", key="status_claim")
    with rc2:
        custom_order_action = st.button("도면참조", key="status_custom_order", help="도면 주고 제작 요청 등")
    with rc3:
        back_to_received_action = st.button("접수", key="status_received")
    with rc4:
        close_action = st.button("마감", key="status_close")
    with rc5:
        cancel_action = st.button("납품취소", key="status_cancel")
    with rc5:
        st.download_button(
            "로젠 업로드(.xlsx)",
            data=_build_lozen_xlsx_bytes_for_orders(orders, items_all, pick_ids) if pick_ids else b"",
            file_name=f"lozen_upload_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            disabled=not bool(pick_ids),
            help="선택된 주문 중 택배(엑셀 기준)만 포함해 로젠 업로드용 엑셀을 만듭니다.",
        )

    # 도면참조/도면참고 주문은 출고 시 금액(단가) 수기 입력을 지원한다.
    if "mutomo_pending_ship_with_price" not in st.session_state:
        st.session_state["mutomo_pending_ship_with_price"] = None

    if ship_action and pick_ids:
        sub = orders[orders["order_id"].astype(str).isin([str(x) for x in pick_ids])]
        if "status" not in sub.columns:
            sub_st = pd.Series(["접수"] * len(sub))
        else:
            sub_st = sub["status"].astype(str).str.strip()
        already_shipped = sub[sub_st == "출고"]["order_id"].astype(str).tolist()
        pending_ship = [str(x) for x in pick_ids if str(x) not in set(already_shipped)]
        if not pending_ship:
            st.warning(
                "선택한 주문이 모두 **이미 출고 처리**되었습니다. "
                "미출고(접수·도면참조·클레임 등) 주문만 골라 주세요."
            )
        else:
            # drawing-ref 여부 판단(전체 주문 + 품목까지 포함)
            draw_ids: list[str] = []
            if "order_id" in orders_all.columns and pending_ship:
                sub_all = orders_all[orders_all["order_id"].astype(str).isin([str(x) for x in pending_ship])].copy()
                for _, rr in sub_all.iterrows():
                    if _is_drawing_ref_order_row(rr, items):
                        draw_ids.append(str(rr.get("order_id") or "").strip())

            if draw_ids:
                st.session_state["mutomo_pending_ship_with_price"] = {
                    "ids": pending_ship,
                    "draw_ids": draw_ids,
                }
                st.rerun()

            if already_shipped:
                st.info(f"이미 출고 {len(already_shipped)}건은 건너뛰고, **{len(pending_ship)}건**만 출고 처리합니다.")
            con = sqlite3.connect(db_path)
            try:
                cur = con.cursor()
                now = dt.datetime.now().isoformat(timespec="seconds")
                cur.executemany(
                    "UPDATE orders SET status=?, shipped_at=COALESCE(NULLIF(shipped_at,''), ?), closed_at=NULL WHERE order_id=?",
                    [("출고", now, oid) for oid in pending_ship],
                )
                con.commit()
            finally:
                con.close()
            st.success("출고 처리했습니다.")
            st.cache_data.clear()
            st.rerun()

    pend_ship = st.session_state.get("mutomo_pending_ship_with_price")
    if isinstance(pend_ship, dict):
        p_ids = [str(x) for x in (pend_ship.get("ids") or []) if str(x)]
        draw_ids = [str(x) for x in (pend_ship.get("draw_ids") or []) if str(x)]
        if p_ids and draw_ids:
            st.warning("도면참고/도면참조 주문이 포함되어 **출고 시 금액(단가) 입력**이 필요합니다.")
            sub_all = orders_all[orders_all["order_id"].astype(str).isin(p_ids)].copy() if "order_id" in orders_all.columns else pd.DataFrame()
            overrides = _load_order_amount_overrides(db_path, set(draw_ids))
            with st.form("ship_drawing_ref_price_form", clear_on_submit=False):
                st.caption("도면참고/도면참조 주문만 입력합니다. (원하면 0으로 둘 수도 있습니다.)")
                for oid in draw_ids:
                    row = sub_all[sub_all["order_id"].astype(str) == oid].head(1)
                    nm = ""
                    if len(row) and "receiver_name" in row.columns:
                        nm = str(row.iloc[0].get("receiver_name") or "").strip()
                    prev_sale, prev_gj = overrides.get(oid, (0.0, 0.0))
                    c1, c2, c3 = st.columns([2.5, 1.2, 1.2])
                    with c1:
                        st.markdown(f"**{nm or '(이름없음)'}**  \n`{oid}`")
                    with c2:
                        st.number_input(
                            "판매금액(총)",
                            min_value=0.0,
                            step=1000.0,
                            value=float(prev_sale),
                            key=f"ship_override_sale_{oid}",
                        )
                    with c3:
                        st.number_input(
                            "광진금액(총)",
                            min_value=0.0,
                            step=1000.0,
                            value=float(prev_gj),
                            key=f"ship_override_gj_{oid}",
                        )
                    st.text_input("메모", value="", key=f"ship_override_note_{oid}", placeholder="예: 도면참고 수기 단가")
                    st.divider()

                c_can, c_ok = st.columns(2)
                with c_can:
                    cancel = st.form_submit_button("취소", type="secondary")
                with c_ok:
                    submit = st.form_submit_button("입력 후 출고 처리", type="primary")

            if cancel:
                st.session_state["mutomo_pending_ship_with_price"] = None
                st.rerun()
            if submit:
                # save overrides for draw ids
                for oid in draw_ids:
                    sale_v = float(st.session_state.get(f"ship_override_sale_{oid}") or 0.0)
                    gj_v = float(st.session_state.get(f"ship_override_gj_{oid}") or 0.0)
                    note_v = str(st.session_state.get(f"ship_override_note_{oid}") or "").strip()
                    _upsert_order_amount_override(db_path, oid, sale_v, gj_v, note_v)

                con = sqlite3.connect(db_path)
                try:
                    cur = con.cursor()
                    now = dt.datetime.now().isoformat(timespec="seconds")
                    cur.executemany(
                        "UPDATE orders SET status=?, shipped_at=COALESCE(NULLIF(shipped_at,''), ?), closed_at=NULL WHERE order_id=?",
                        [("출고", now, oid) for oid in p_ids],
                    )
                    con.commit()
                finally:
                    con.close()
                st.session_state["mutomo_pending_ship_with_price"] = None
                st.success("금액 저장 후 출고 처리했습니다.")
                st.cache_data.clear()
                st.rerun()

    if claim_action and pick_ids:
        con = sqlite3.connect(db_path)
        try:
            cur = con.cursor()
            cur.executemany(
                "UPDATE orders SET status=?, closed_at=NULL WHERE order_id=?",
                [("클레임", oid) for oid in pick_ids],
            )
            con.commit()
        finally:
            con.close()
        st.success("클레임으로 변경했습니다.")
        st.cache_data.clear()
        st.rerun()

    if custom_order_action and pick_ids:
        con = sqlite3.connect(db_path)
        try:
            cur = con.cursor()
            cur.executemany(
                "UPDATE orders SET status=?, closed_at=NULL WHERE order_id=?",
                [("도면참조", oid) for oid in pick_ids],
            )
            con.commit()
        finally:
            con.close()
        st.success("도면참조로 변경했습니다.")
        st.cache_data.clear()
        st.rerun()

    if back_to_received_action and pick_ids:
        if "order_id" in orders_all.columns:
            sub0 = orders_all[orders_all["order_id"].astype(str).isin([str(x) for x in pick_ids])].copy()
        else:
            sub0 = pd.DataFrame()
        need_warn, wmsg = _revert_to_received_warn_message(sub0)
        if need_warn:
            st.session_state[MUTOMO_PENDING_REVERT_KEY] = {
                "order_ids": [str(x) for x in pick_ids],
                "source": "search",
                "message": wmsg,
            }
            st.rerun()
        con = sqlite3.connect(db_path)
        try:
            cur = con.cursor()
            cur.executemany(
                "UPDATE orders SET status=?, shipped_at=NULL, closed_at=NULL WHERE order_id=?",
                [("접수", oid) for oid in pick_ids],
            )
            con.commit()
        finally:
            con.close()
        st.success("접수로 변경했습니다. (출고시간/마감시간 초기화)")
        st.cache_data.clear()
        st.rerun()

    if close_action and pick_ids:
        con = sqlite3.connect(db_path)
        try:
            cur = con.cursor()
            cur.executemany(
                "UPDATE orders SET status=?, closed_at=COALESCE(closed_at, ?) WHERE order_id=?",
                [("마감", dt.datetime.now().isoformat(timespec="seconds"), oid) for oid in pick_ids],
            )
            con.commit()
        finally:
            con.close()
        st.success("마감으로 변경했습니다.")
        st.cache_data.clear()
        st.rerun()

    if cancel_action and pick_ids:
        con = sqlite3.connect(db_path)
        try:
            cur = con.cursor()
            cur.executemany(
                "UPDATE orders SET status=?, closed_at=NULL WHERE order_id=?",
                [("납품취소", oid) for oid in pick_ids],
            )
            con.commit()
        finally:
            con.close()
        st.success("납품취소로 변경했습니다.")
        st.cache_data.clear()
        st.rerun()

    sig_key = "|".join(parts_live)

    if parts_live and len(hits_live):
        if "order_id" in hits_live.columns:
            hit_id_set = set(hits_live["order_id"].astype(str).tolist())
            if pick_ids:
                oids = [str(x) for x in pick_ids if str(x) in hit_id_set]
            else:
                oids = []
            if len(oids) > _MAX_SEARCH_TO_TABLE_ROWS:
                st.caption(
                    f"선택 {len(oids)}건 중 표·사이드바에는 앞에서 {_MAX_SEARCH_TO_TABLE_ROWS}건만 연동합니다. "
                    "더 좁히려면 오른쪽에서 선택을 줄이거나 검색어를 나누세요."
                )
                oids = oids[:_MAX_SEARCH_TO_TABLE_ROWS]
            st.session_state["_search_sync_order_ids"] = oids
            st.session_state["table_sync_from_search"] = True
            st.session_state["selected_ids_from_table"] = list(oids)
            st.session_state["prev_table_pick_ids"] = list(oids)
            st.session_state["active_selector"] = "table"
            sig = (sig_key, tuple(oids))
            if st.session_state.get("_search_sync_sig") != sig:
                st.session_state["_search_sync_sig"] = sig
                st.session_state["_search_sync_need_df_apply"] = True
    elif parts_live and len(hits_live) == 0:
        _ts0 = st.session_state.get("table_sync_from_search")
        st.session_state.pop("search_pick_labels", None)
        st.session_state["search_pick_ids"] = []
        st.session_state["prev_search_pick_ids"] = []
        if _ts0:
            st.session_state["selected_ids_from_table"] = []
            st.session_state["prev_table_pick_ids"] = []
            st.session_state["active_selector"] = None
        cand = _unique_receiver_names(nh)
        merged: list[str] = []
        for p in parts_live:
            merged.extend(_fuzzy_receiver_name_suggestions(p, cand))
        seen: set[str] = set()
        sugs: list[str] = []
        for s in merged:
            if s not in seen:
                seen.add(s)
                sugs.append(s)
        sugs = sugs[:12]
        if sugs:
            st.caption("등록된 이름 중 비슷한 후보. 누르면 **OR 표에 추가**됩니다.")
            ncols = min(4, len(sugs))
            for row_start in range(0, len(sugs), ncols):
                row = sugs[row_start : row_start + ncols]
                cols = st.columns(len(row))
                for j, name in enumerate(row):
                    idx = row_start + j
                    if cols[j].button(name, key=f"_name_suggest_btn_{idx}"):
                        if name not in st.session_state["name_search_terms"]:
                            st.session_state["name_search_terms"].append(name)
                        st.rerun()
        st.session_state.pop("_search_sync_order_ids", None)
        st.session_state.pop("_search_sync_sig", None)
        st.session_state["table_sync_from_search"] = False
        if _ts0 and "mutomo_full_orders_df" in st.session_state:
            st.session_state["mutomo_full_orders_df"] = {
                "selection": {"rows": [], "columns": [], "cells": []},
            }
    elif not parts_live:
        _full_reset = st.session_state.pop("_search_ui_full_reset", False)
        _ts_empty = st.session_state.get("table_sync_from_search")
        st.session_state["search_pick_ids"] = []
        st.session_state["prev_search_pick_ids"] = []
        st.session_state.pop("search_pick_labels", None)
        if _full_reset or _ts_empty:
            st.session_state["selected_ids_from_table"] = []
            st.session_state["prev_table_pick_ids"] = []
            st.session_state["active_selector"] = None
        st.session_state.pop("_search_sync_order_ids", None)
        st.session_state.pop("_search_sync_sig", None)
        st.session_state["table_sync_from_search"] = False
        if (_full_reset or _ts_empty) and "mutomo_full_orders_df" in st.session_state:
            st.session_state["mutomo_full_orders_df"] = {
                "selection": {"rows": [], "columns": [], "cells": []},
            }

    if st.session_state.pop("_focus_next_name_inline", False):
        _focus_streamlit_input_by_placeholder("다음 이름")


def render_sales_period_tab(orders_all: pd.DataFrame, items_all: pd.DataFrame) -> None:
    """3번째 탭: 마감(청구) 전용 화면."""
    st.subheader("마감(청구) 관리")
    st.caption("복잡한 집계는 숨기고, **마감/입금/미수/이슈**만 관리합니다.")
    today = dt.date.today()
    c1, c2 = st.columns(2)
    with c1:
        d_start = st.date_input("기간 시작일(출고일)", value=today - dt.timedelta(days=30), key="dash_sales_d0")
    with c2:
        d_end = st.date_input("기간 종료일(출고일)", value=today, key="dash_sales_d1")
    if d_start > d_end:
        st.warning("시작일이 종료일보다 늦습니다.")
        return

    _root = os.path.dirname(os.path.abspath(__file__))
    price_path = os.path.join(_root, "단가표.csv")
    price_mtime = os.path.getmtime(price_path) if os.path.isfile(price_path) else -1.0
    price_map, price_warns = _price_map_cached(price_path, float(price_mtime))

    if price_warns:
        st.info("단가표: " + " ".join(w for w in price_warns if w))

    st.divider()
    st.caption("기준: 출고일(`shipped_at`, 한국시간). shipped_at이 없는 주문은 기간에 포함되지 않습니다.")

    k0, k1 = st.columns(2)
    with k0:
        close_d0 = st.date_input("마감 대상 시작일(출고일)", value=d_start, key="dash_close_d0")
    with k1:
        close_d1 = st.date_input("마감 대상 종료일(출고일)", value=d_end, key="dash_close_d1")

    if close_d0 > close_d1:
        st.warning("시작일이 종료일보다 늦습니다.")
    else:
        db_path = str(st.session_state.get("mutomo_db_path", "mutomo.sqlite"))
        _ensure_settlements_table(db_path)

        base = orders_all.copy()
        if "shipped_at" in base.columns:
            base["_ship_day"] = _shipped_at_calendar_date_series(base["shipped_at"])
        else:
            base["_ship_day"] = pd.Series([None] * len(base), index=base.index, dtype=object)

        in_ship_range = base["_ship_day"].notna() & (base["_ship_day"] >= close_d0) & (base["_ship_day"] <= close_d1)
        base = base.loc[in_ship_range].copy()

        stt = base["status"].astype(str).str.strip() if "status" in base.columns else pd.Series([""] * len(base))
        is_closed = pd.Series([False] * len(base), index=base.index)
        if "closed_at" in base.columns:
            ca = base["closed_at"].astype(str).fillna("").str.strip()
            is_closed = (ca != "") & (ca.str.lower() != "nan")

        # 마감 대상(기본): 출고 상태 & 아직 closed_at 없음
        cand = base[(stt == "출고") & (~is_closed)].copy()
        already = base[(stt == "마감") | is_closed].copy()

        # 금액 요약 (단가표 기준)
        closed_ids = set(already["order_id"].astype(str).tolist()) if "order_id" in already.columns else set()
        cand_ids = set(cand["order_id"].astype(str).tolist()) if "order_id" in cand.columns else set()
        ov = _load_order_amount_overrides(db_path, set(closed_ids) | set(cand_ids))
        closed_sale, closed_gj, closed_unp_lines, closed_unp_qty = _calc_amounts_for_orders(
            items_all, closed_ids, price_map, overrides=ov
        )
        cand_sale, cand_gj, cand_unp_lines, cand_unp_qty = _calc_amounts_for_orders(
            items_all, cand_ids, price_map, overrides=ov
        )

        # 배송 유형(택배/직접/혼합)별 금액/건수
        def _bucket_sums(order_ids: set[str]) -> dict[str, dict[str, float | int]]:
            d: dict[str, dict[str, float | int]] = {k: {"n": 0, "sale": 0.0, "gj": 0.0} for k in ("택배", "직접", "혼합")}
            if not order_ids:
                return d
            amt = _calc_order_amount_map(items_all, order_ids, price_map, overrides=ov)
            for oid in order_ids:
                b = _ship_bucket_for_order(items_all, str(oid))
                sale, gj = amt.get(str(oid), (0.0, 0.0))
                d[b]["n"] = int(d[b]["n"]) + 1
                d[b]["sale"] = float(d[b]["sale"]) + float(sale)
                d[b]["gj"] = float(d[b]["gj"]) + float(gj)
            return d

        closed_by = _bucket_sums(closed_ids)
        cand_by = _bucket_sums(cand_ids)

        m_a, m_b, m_c = st.columns(3)
        with m_a:
            st.metric("기간 내 출고(전체)", int((stt == "출고").sum()))
        with m_b:
            st.metric("마감 대상(출고 & 미마감)", int(len(cand)))
        with m_c:
            st.metric("이미 마감(또는 closed_at 존재)", int(len(already)))

        a1, a2, a3 = st.columns(3)
        with a1:
            st.metric("마감된 금액(판매)", format_won(closed_sale))
        with a2:
            st.metric("마감된 금액(광진)", format_won(closed_gj))
        with a3:
            st.metric("미마감 출고 금액(판매)", format_won(cand_sale))
        if closed_unp_qty or cand_unp_qty:
            st.caption(
                f"단가 미매칭(참고) — 마감: {closed_unp_lines}라인/{closed_unp_qty}개, "
                f"미마감: {cand_unp_lines}라인/{cand_unp_qty}개"
            )

        # 입금/미수 관리 (기간 단위)
        period_key = (close_d0.isoformat(), close_d1.isoformat(), "shipped")
        paid_default = 0.0
        note_default = ""
        try:
            con_s = sqlite3.connect(db_path)
            try:
                row = con_s.execute(
                    "SELECT paid_amount, note FROM settlements WHERE period_start=? AND period_end=? AND basis=?",
                    period_key,
                ).fetchone()
                if row:
                    paid_default = float(row[0] or 0)
                    note_default = "" if row[1] is None else str(row[1])
            finally:
                con_s.close()
        except Exception:
            pass

        p1, p2, p3 = st.columns(3)
        with p1:
            paid_amount = st.number_input(
                "입금된 금액(판매 기준)",
                min_value=0.0,
                step=10000.0,
                value=float(paid_default),
                key="dash_paid_amount",
                help="이 기간 청구분 중 실제 입금 확인 금액을 적어두면 미수금액이 자동 계산됩니다.",
            )
        with p2:
            st.metric("미수 금액(판매 기준)", format_won(max(0.0, closed_sale - float(paid_amount))))
        with p3:
            st.metric("참고: 마감금액-입금금액", format_won(float(closed_sale) - float(paid_amount)))
        note = st.text_input("메모(선택)", value=note_default, key="dash_settlement_note", placeholder="예) 4/21 입금 확인, 세금계산서 발행완료")
        if st.button("입금/메모 저장", key="dash_settlement_save"):
            now = dt.datetime.now().isoformat(timespec="seconds")
            con_s2 = sqlite3.connect(db_path)
            try:
                con_s2.execute(
                    """
                    INSERT INTO settlements(period_start, period_end, basis, expected_sale, expected_gj, paid_amount, note, updated_at)
                    VALUES(?,?,?,?,?,?,?,?)
                    ON CONFLICT(period_start, period_end, basis) DO UPDATE SET
                      expected_sale=excluded.expected_sale,
                      expected_gj=excluded.expected_gj,
                      paid_amount=excluded.paid_amount,
                      note=excluded.note,
                      updated_at=excluded.updated_at
                    """,
                    (
                        period_key[0],
                        period_key[1],
                        period_key[2],
                        float(closed_sale),
                        float(closed_gj),
                        float(paid_amount),
                        note,
                        now,
                    ),
                )
                con_s2.commit()
            finally:
                con_s2.close()
            st.success("저장했습니다.")

        st.subheader("배송유형별(택배/직접/혼합) 금액")
        rows = []
        for b in ("택배", "직접", "혼합"):
            rows.append(
                {
                    "구분": b,
                    "미수(출고)건수": int(cand_by[b]["n"]),
                    "미수(출고)판매": format_won(cand_by[b]["sale"]),
                    "마감건수": int(closed_by[b]["n"]),
                    "마감판매": format_won(closed_by[b]["sale"]),
                }
            )
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True, height=160)

        if len(cand):
            prev_cols = [c for c in ["_ship_day", "receiver_name", "order_list", "order_id"] if c in cand.columns]
            prev = cand[prev_cols].copy()
            if "order_list" in prev.columns:
                prev["order_list"] = prev["order_list"].astype(str).fillna("").map(lambda x: (x[:80] + "…") if len(x) > 81 else x)
            prev = prev.rename(columns={"_ship_day": "출고일", "receiver_name": "받는분"})
            st.dataframe(prev.head(80), use_container_width=True, hide_index=True, height=260)

            confirm = st.checkbox(
                f"{close_d0.isoformat()} ~ {close_d1.isoformat()} 출고 {len(cand)}건을 마감 처리",
                value=False,
                key="dash_close_confirm",
            )
            if st.button("기간 마감 실행", type="primary", disabled=(not confirm), key="dash_close_run"):
                oids = cand["order_id"].astype(str).tolist() if "order_id" in cand.columns else []
                if not oids:
                    st.warning("order_id가 없어 처리할 수 없습니다.")
                else:
                    now = dt.datetime.now().isoformat(timespec="seconds")
                    con = sqlite3.connect(st.session_state.get("mutomo_db_path", "mutomo.sqlite"))
                    try:
                        cur = con.cursor()
                        cur.executemany(
                            "UPDATE orders SET status=?, closed_at=COALESCE(closed_at, ?) WHERE order_id=?",
                            [("마감", now, oid) for oid in oids],
                        )
                        con.commit()
                    finally:
                        con.close()
                    st.success(f"기간 마감 완료: {len(oids)}건")
                    st.cache_data.clear()
                    st.rerun()
        else:
            st.caption("이 기간에 **마감할 출고 주문**이 없습니다.")

        st.subheader("최근 마감 내역")
        try:
            con_r = sqlite3.connect(st.session_state.get("mutomo_db_path", "mutomo.sqlite"))
            try:
                recent_closed = pd.read_sql_query(
                    """
                    select
                      closed_at as 마감시각,
                      shipped_at as 출고시각,
                      receiver_name as 받는분,
                      order_id as order_id
                    from orders
                    where status = '마감' and closed_at is not null and trim(closed_at) <> ''
                    order by closed_at desc
                    limit 30
                    """,
                    con_r,
                )
            finally:
                con_r.close()
            if len(recent_closed):
                st.dataframe(recent_closed, use_container_width=True, hide_index=True, height=240)
            else:
                st.caption("최근 마감 내역이 없습니다.")
        except Exception:
            st.caption("마감 내역을 불러오지 못했습니다. 새로고침 후 다시 확인해 주세요.")

        st.subheader("기간 이슈(클레임/취소/메모)")
        # 기간 내 출고(또는 마감 포함) 중 이슈만
        has_issue_status = stt.isin({"클레임", "납품취소"})
        has_memo = pd.Series([False] * len(base), index=base.index)
        if "special_issue" in base.columns:
            si = base["special_issue"].astype(str).fillna("").str.strip()
            has_memo = has_memo | ((si != "") & (si.str.lower() != "nan"))
        if "attention_note" in base.columns:
            an = base["attention_note"].astype(str).fillna("").str.strip()
            has_memo = has_memo | ((an != "") & (an.str.lower() != "nan"))
        issues = base[has_issue_status | has_memo].copy()
        st.metric("이슈 건수", int(len(issues)))
        if len(issues):
            cols = [c for c in ["_ship_day", "receiver_name", "status", "special_issue", "attention_note", "order_id"] if c in issues.columns]
            view = issues[cols].copy()
            view = view.rename(
                columns={
                    "_ship_day": "출고일",
                    "receiver_name": "받는분",
                    "status": "상태",
                    "special_issue": "특이사항(수기)",
                    "attention_note": "특이사항(엑셀)",
                    "order_id": "order_id",
                }
            )
            for c in ["특이사항(수기)", "특이사항(엑셀)"]:
                if c in view.columns:
                    view[c] = view[c].astype(str).fillna("").map(lambda x: (x[:80] + "…") if len(x) > 81 else x)
            st.dataframe(view, use_container_width=True, hide_index=True, height=300)
        else:
            st.caption("이 기간에는 클레임/납품취소/특이사항이 없습니다.")

        with st.expander("마감 리스트(기간 내) 펼쳐보기", expanded=False):
            if len(already) == 0:
                st.caption("이 기간에 마감된 주문이 없습니다.")
            else:
                max_rows = 2000
                show_df = already.copy()
                if len(show_df) > max_rows:
                    st.caption(f"마감 리스트는 상위 {max_rows:,}건만 표시합니다. (현재 {len(show_df):,}건)")
                    show_df = show_df.head(max_rows)
                oids = set(show_df["order_id"].astype(str).tolist()) if "order_id" in show_df.columns else set()
                amt_map = _calc_order_amount_map(items_all, oids, price_map, overrides=ov)
                show_df["_판매금액"] = show_df["order_id"].astype(str).map(lambda x: amt_map.get(str(x), (0.0, 0.0))[0])
                show_df["_배송유형"] = show_df["order_id"].astype(str).map(lambda x: _ship_bucket_for_order(items_all, str(x)))
                cols = [c for c in ["_ship_day", "closed_at", "receiver_name", "_배송유형", "_판매금액", "special_issue", "attention_note", "order_id"] if c in show_df.columns]
                view = show_df[cols].copy()
                view = view.rename(
                    columns={
                        "_ship_day": "출고일",
                        "closed_at": "마감시각",
                        "receiver_name": "받는분",
                        "_배송유형": "배송",
                        "_판매금액": "판매금액(단가표)",
                        "special_issue": "특이사항(수기)",
                        "attention_note": "특이사항(엑셀)",
                        "order_id": "order_id",
                    }
                )
                if "판매금액(단가표)" in view.columns:
                    view["판매금액(단가표)"] = view["판매금액(단가표)"].map(format_won)
                for c in ["특이사항(수기)", "특이사항(엑셀)"]:
                    if c in view.columns:
                        view[c] = view[c].astype(str).fillna("").map(lambda x: (x[:80] + "…") if len(x) > 81 else x)
                st.dataframe(view, use_container_width=True, hide_index=True, height=420)


def main() -> None:
    st.set_page_config(page_title="Mutomo 판매관리", layout="wide")

    # Wider sidebar + 메인 본문을 위로(기본 상단 패딩 축소)
    st.markdown(
        """
<style>
  /* Sidebar width: keep wide when expanded, but let it fully collapse without "ghost" space */
  section[data-testid="stSidebar"] { width: 560px !important; min-width: 560px !important; }
  section[data-testid="stSidebar"] > div { width: 560px !important; min-width: 560px !important; }
  section[data-testid="stSidebar"][aria-expanded="false"] { width: 0px !important; min-width: 0px !important; }
  section[data-testid="stSidebar"][aria-expanded="false"] > div { width: 0px !important; min-width: 0px !important; }
  .block-container { padding-top: 0.75rem !important; padding-bottom: 0.5rem !important; }
</style>
        """,
        unsafe_allow_html=True,
    )

    init_shared_session_state()

    # DB missing / unreadable messages (written from sidebar 설정 expander)
    main_db_slot = st.empty()

    # Sidebar top navigation (replaces main top tabs/links)
    if "mutomo_page" not in st.session_state:
        st.session_state["mutomo_page"] = "page1"
    with st.sidebar.container():
        c_nav1, c_nav2, c_nav3 = st.columns(3)
        cur_page = str(st.session_state.get("mutomo_page") or "page1")

        def _go(page_key: str) -> None:
            st.session_state["mutomo_page"] = page_key
            st.rerun()

        with c_nav1:
            if st.button("출고", type=("primary" if cur_page == "page2" else "secondary"), key="nav_page2"):
                _go("page2")
        with c_nav2:
            if st.button("접수목록", type=("primary" if cur_page == "page1" else "secondary"), key="nav_page1"):
                _go("page1")
        with c_nav3:
            if st.button("마감", type=("primary" if cur_page == "page3" else "secondary"), key="nav_page3"):
                _go("page3")
        st.divider()

    # Settings in sidebar (collapsed)
    with st.sidebar.expander("설정", expanded=False):
        db_path = st.text_input("DB 경로", value="mutomo.sqlite", key="mutomo_db_path")
        q1, q2, q3 = st.columns(3)
        with q1:
            if st.button("전체", key="quick_status_all", help="상태 필터를 전체로"):
                st.session_state["mutomo_status_filter"] = list(ORDER_STATUS_FILTER_OPTIONS)
                st.rerun()
        with q2:
            if st.button("마감", key="quick_status_closed", help="마감만 보기"):
                st.session_state["mutomo_status_filter"] = ["마감"]
                st.rerun()
        with q3:
            if st.button("클레임", key="quick_status_claim", help="클레임만 보기"):
                st.session_state["mutomo_status_filter"] = ["클레임"]
                st.rerun()
        status_filter = st.multiselect(
            "상태 필터",
            options=list(ORDER_STATUS_FILTER_OPTIONS),
            default=list(ORDER_STATUS_FILTER_OPTIONS),
            key="mutomo_status_filter",
        )
        date_basis = st.selectbox(
            "‘오늘 접수’ 기준",
            options=["purchase_date(구매일자: 파일명)", "created_at(수집시각)"],
            index=0,
        )
        show_admin_cols = st.toggle("관리 컬럼 보기", value=False)
        try:
            mtime = dt.datetime.fromtimestamp(os.path.getmtime(__file__)).isoformat(timespec="seconds")
        except Exception:
            mtime = "unknown"
        st.caption(f"dashboard.py: {__file__} (mtime: {mtime})")

        # 자동 DB 백업: 로컬 하루 최대 4회(6시간마다 한 번), 30일 넘은 백업 파일은 삭제
        _prune_sqlite_backups(db_path, backup_dir="backups", keep_days=BACKUP_RETENTION_DAYS)
        if not _sqlite_slot_already_has_backup(db_path, "backups"):
            out = _backup_sqlite(db_path, backup_dir="backups", keep_days=BACKUP_RETENTION_DAYS)
            if out:
                st.caption(f"DB 자동백업(이번 6시간 구간): `{out}`")

        st.caption(
            "**엑셀→DB 수집**: 이 프로젝트의 `order_list` 폴더에 있는 `.xlsx`를 읽어 아래 DB 경로의 SQLite에 반영합니다. "
            "**DB 다시 읽기**: 디스크에 이미 반영된 DB만 화면 캐시에서 다시 불러옵니다."
        )
        _ing1, _ing2 = st.columns(2)
        with _ing1:
            if st.button(
                "엑셀→DB 수집",
                key="ingest_order_list_btn",
                help="`order_list` 전체 .xlsx → 설정의 DB 파일(터미널 `ingest_xlsx.py`와 동일)",
                use_container_width=True,
            ):
                with st.spinner("`order_list`에서 엑셀 읽는 중…"):
                    _ok_ing, _msg_ing = run_ingest(
                        db_path,
                        input_dir="order_list",
                        aliases_path="product_aliases.yml",
                    )
                if _ok_ing:
                    st.session_state["mutomo_db_reload_nonce"] = int(st.session_state.get("mutomo_db_reload_nonce", 0)) + 1
                    st.session_state["mutomo_toast_after_ingest"] = _msg_ing
                    try:
                        load_tables.clear()
                    except Exception:
                        pass
                    st.cache_data.clear()
                    st.rerun()
                else:
                    st.error(_msg_ing)
        with _ing2:
            if st.button(
                "DB 다시 읽기",
                key="reload_db_cache",
                help="DB 파일은 그대로 두고 화면만 최신 DB로 갱신",
                use_container_width=True,
            ):
                st.session_state["mutomo_db_reload_nonce"] = int(st.session_state.get("mutomo_db_reload_nonce", 0)) + 1
                st.session_state["mutomo_show_db_reload_toast"] = True
                try:
                    load_tables.clear()
                except Exception:
                    pass
                st.cache_data.clear()
                st.rerun()

        ok_db, db_msg = _db_ready(db_path)
        if not ok_db:
            main_db_slot.error(db_msg)
            main_db_slot.markdown(
                "아래 명령으로 `orders` / `items` 테이블을 만듭니다. 기본으로 **`order_list` 폴더** 안의 `.xlsx`만 읽습니다. "
                "그 폴더에 파일이 없으면 **빈 테이블만** 만들어져 대시보드가 열리고, 엑셀을 넣고 같은 명령을 다시 실행하면 데이터가 채워집니다.\n\n"
                "```bash\n"
                "python ingest_xlsx.py --db mutomo.sqlite --aliases product_aliases.yml\n"
                "```\n\n"
                "다른 위치에만 두고 싶으면 `--input-dir \"D:\\경로\"` 처럼 지정하면 됩니다. 다른 이름의 DB를 쓰는 경우 **설정**의 DB 경로를 그 파일로 맞추세요."
            )
            st.stop()

        _migrate_orders_schema(db_path)
        _mt, _sz = _db_stat_for_cache(db_path)
        _rn = int(st.session_state.get("mutomo_db_reload_nonce", 0))
        orders_all, items_all = load_tables(db_path, _mt, _sz, _rn)
        if _ti := st.session_state.pop("mutomo_toast_after_ingest", None):
            st.toast(str(_ti), icon="✅")
        elif st.session_state.pop("mutomo_show_db_reload_toast", False):
            st.toast(
                f"DB를 다시 읽었습니다. 주문 {len(orders_all):,}건, 품목 {len(items_all):,}줄.",
                icon="✅",
            )

        try:
            con0 = sqlite3.connect(db_path)
            try:
                _lim = int(_RECENT_SOURCE_FILES_SIDEBAR_PREVIEW)
                recent_files = pd.read_sql_query(
                    f"""
                    select
                      source_file as 파일,
                      count(*) as 주문건수,
                      max(created_at) as 마지막수집
                    from orders
                    group by source_file
                    order by 마지막수집 desc
                    limit {_lim}
                    """,
                    con0,
                )
            finally:
                con0.close()
            if len(recent_files):
                st.markdown(f"**최근 수집 소스 파일** (마지막 수집 시각 기준 **이 표에만** 상위 {_lim}건)")
                st.caption(
                    "ingest를 한 달에 한 번이든 매일이든 상관없습니다. DB에는 그대로 쌓이고, "
                    f"여기는 사이드바 미리보기라 **상위 {_lim}건만** 보여 줍니다. 나머지는 표 안에서 스크롤하세요."
                )
                st.dataframe(recent_files, use_container_width=True, hide_index=True, height=220)
        except Exception:
            pass

    with st.sidebar.expander("출고 목록 엑셀", expanded=False):
        export_version = "v12_strip_order_list_from_notes"
        shipped_date = st.date_input("출고 기준 날짜", value=dt.date.today(), key="export_shipped_date")
        shipped_today_cnt = 0
        if "status" in orders_all.columns:
            shipped = orders_all[orders_all["status"] == "출고"].copy()
            if "shipped_at" in shipped.columns:
                shipped["_d"] = _shipped_at_calendar_date_series(shipped["shipped_at"])
                shipped_today_cnt = int((shipped["_d"] == shipped_date).sum())
            else:
                shipped_today_cnt = len(shipped)
        st.caption(f"{shipped_date.isoformat()} 출고: {shipped_today_cnt}건")
        if shipped_today_cnt > 0 and "status" in orders_all.columns:
            shipped_pick = orders_all[orders_all["status"] == "출고"].copy()
            if "shipped_at" in shipped_pick.columns:
                shipped_pick["_d"] = _shipped_at_calendar_date_series(shipped_pick["shipped_at"])
                shipped_pick = shipped_pick[shipped_pick["_d"] == shipped_date]
            _nu, _nt, _nd, _nm = _picking_stats(shipped_pick, items_all)
            st.caption(f"피킹 시트 행: 택배 {_nt}, 직접 {_nd} — 혼합 {_nm}건은 두 시트에 각 1행 (검산 {_nt}+{_nd} = {_nu}+{_nm})")
        filename = f"mutomo_shipped_{shipped_date.isoformat()}_shipped.xlsx"
        shipped_today_ids = []
        if shipped_today_cnt > 0 and "status" in orders_all.columns:
            shipped = orders_all[orders_all["status"] == "출고"].copy()
            if "shipped_at" in shipped.columns:
                shipped["_d"] = _shipped_at_calendar_date_series(shipped["shipped_at"])
                shipped = shipped[shipped["_d"] == shipped_date]
            shipped_today_ids = shipped.get("order_id", pd.Series([], dtype=str)).astype(str).tolist()
        shipped_key = f"{shipped_date.isoformat()}|{shipped_today_cnt}|{'/'.join(shipped_today_ids[:200])}"
        xlsx_bytes = _today_shipped_excel_cached(export_version, shipped_key, orders_all, items_all) if shipped_today_cnt > 0 else b""

        st.download_button(
            "출고 목록 엑셀 다운로드",
            data=xlsx_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            disabled=(shipped_today_cnt == 0),
            key=f"download_shipped_{export_version}_{shipped_date.isoformat()}_{shipped_today_cnt}",
        )

    with st.sidebar.expander("기간별 마감(청구용)", expanded=False):
        st.caption("출고일(shipped_at, 한국시간) 기준으로 기간 내 주문을 한 번에 **마감** 처리합니다.")
        d0, d1 = st.columns(2)
        with d0:
            close_d0 = st.date_input("시작일", value=dt.date.today() - dt.timedelta(days=7), key="close_period_d0")
        with d1:
            close_d1 = st.date_input("종료일", value=dt.date.today(), key="close_period_d1")
        if close_d0 > close_d1:
            st.warning("시작일이 종료일보다 늦습니다.")
        else:
            close_statuses = st.multiselect(
                "대상 상태",
                options=list(ORDER_STATUS_FILTER_OPTIONS),
                default=["출고"],
                key="close_period_statuses",
                help="보통은 출고만 선택합니다. 이미 마감인 주문을 포함하면 중복 업데이트가 늘어납니다.",
            )
            st.caption("※ shipped_at이 없는 주문은 기간 필터에 걸리지 않습니다.")
            cand = orders_all.copy()
            if "shipped_at" in cand.columns:
                cand["_ship_day"] = _shipped_at_calendar_date_series(cand["shipped_at"])
            else:
                cand["_ship_day"] = pd.Series([None] * len(cand), index=cand.index, dtype=object)
            mask = cand["_ship_day"].notna() & (cand["_ship_day"] >= close_d0) & (cand["_ship_day"] <= close_d1)
            if close_statuses and "status" in cand.columns:
                mask = mask & cand["status"].astype(str).str.strip().isin(set(close_statuses))
            cand = cand.loc[mask].copy()
            n_cand = len(cand)
            st.metric("마감 대상(건)", n_cand)
            if n_cand:
                prev = cand[["shipped_at", "receiver_name", "status", "order_id"]].copy()
                prev = prev.rename(
                    columns={"shipped_at": "출고처리시각", "receiver_name": "받는분", "status": "상태", "order_id": "order_id"}
                )
                st.dataframe(prev.head(50), use_container_width=True, hide_index=True, height=210)
                confirm = st.checkbox(
                    f"{close_d0.isoformat()} ~ {close_d1.isoformat()} 기간의 {n_cand}건을 마감 처리",
                    value=False,
                    key="close_period_confirm",
                )
                if st.button("기간 마감 실행", type="primary", disabled=(not confirm), key="close_period_run"):
                    oids = cand["order_id"].astype(str).tolist() if "order_id" in cand.columns else []
                    if not oids:
                        st.warning("order_id가 없어 처리할 수 없습니다.")
                    else:
                        con = sqlite3.connect(db_path)
                        try:
                            cur = con.cursor()
                            cur.executemany(
                                "UPDATE orders SET status=?, closed_at=COALESCE(closed_at, ?) WHERE order_id=?",
                                [("마감", dt.datetime.now().isoformat(timespec="seconds"), oid) for oid in oids],
                            )
                            con.commit()
                        finally:
                            con.close()
                        st.success(f"기간 마감 완료: {len(oids)}건")
                        st.cache_data.clear()
                        st.rerun()
            else:
                st.caption("해당 기간·조건에 맞는 주문이 없습니다.")

    with st.sidebar.expander("최근 마감", expanded=False):
        st.caption("마감 처리한 시각(`closed_at`) 기준 최근 내역입니다.")
        try:
            con_r = sqlite3.connect(db_path)
            try:
                recent_closed = pd.read_sql_query(
                    """
                    select
                      closed_at as 마감시각,
                      shipped_at as 출고시각,
                      receiver_name as 받는분,
                      status as 상태,
                      order_id as order_id
                    from orders
                    where status = '마감' and closed_at is not null and trim(closed_at) <> ''
                    order by closed_at desc
                    limit 30
                    """,
                    con_r,
                )
            finally:
                con_r.close()
            if len(recent_closed):
                st.dataframe(recent_closed, use_container_width=True, hide_index=True, height=260)
            else:
                st.caption("아직 마감 내역이 없습니다. (방금 마감했다면 새로고침 후 보입니다)")
        except Exception:
            st.caption("closed_at 컬럼이 아직 없거나(처음 1회), DB 조회에 실패했습니다. 새로고침 후 다시 확인해 주세요.")

    with st.sidebar.expander("기간 이슈(클레임/취소/메모)", expanded=False):
        st.caption("청구(마감) 완료 후에는 여기만 보면 됩니다. 기간 내 출고분에서 **문제될 건**만 모아 보여줍니다.")
        i0, i1 = st.columns(2)
        with i0:
            issue_d0 = st.date_input("시작일", value=dt.date.today() - dt.timedelta(days=7), key="issue_period_d0")
        with i1:
            issue_d1 = st.date_input("종료일", value=dt.date.today(), key="issue_period_d1")
        if issue_d0 > issue_d1:
            st.warning("시작일이 종료일보다 늦습니다.")
        else:
            base = orders_all.copy()
            if "shipped_at" in base.columns:
                base["_ship_day"] = _shipped_at_calendar_date_series(base["shipped_at"])
            else:
                base["_ship_day"] = pd.Series([None] * len(base), index=base.index, dtype=object)
            # 기간 내 출고분만
            m = base["_ship_day"].notna() & (base["_ship_day"] >= issue_d0) & (base["_ship_day"] <= issue_d1)
            if "status" in base.columns:
                m = m & base["status"].astype(str).str.strip().isin({"출고", "마감", "클레임", "납품취소"})
            base = base.loc[m].copy()

            if len(base) == 0:
                st.caption("해당 기간에 출고(shipped_at)된 주문이 없습니다.")
            else:
                stt = base["status"].astype(str).str.strip() if "status" in base.columns else pd.Series([""] * len(base))
                has_issue_status = stt.isin({"클레임", "납품취소"})
                has_memo = pd.Series([False] * len(base), index=base.index)
                if "special_issue" in base.columns:
                    si = base["special_issue"].astype(str).fillna("").str.strip()
                    has_memo = has_memo | (si != "") & (si.str.lower() != "nan")
                if "attention_note" in base.columns:
                    an = base["attention_note"].astype(str).fillna("").str.strip()
                    has_memo = has_memo | (an != "") & (an.str.lower() != "nan")

                issues = base[has_issue_status | has_memo].copy()
                st.metric("이슈 건수", int(len(issues)))
                if len(issues) == 0:
                    st.caption("클레임/납품취소/메모(특이사항)가 없습니다.")
                else:
                    cols = [c for c in ["_ship_day", "receiver_name", "status", "special_issue", "attention_note", "order_id"] if c in issues.columns]
                    view = issues[cols].copy()
                    view = view.rename(
                        columns={
                            "_ship_day": "출고일",
                            "receiver_name": "받는분",
                            "status": "상태",
                            "special_issue": "특이사항(수기)",
                            "attention_note": "특이사항(엑셀)",
                            "order_id": "order_id",
                        }
                    )
                    # 텍스트는 조금만
                    for c in ["특이사항(수기)", "특이사항(엑셀)"]:
                        if c in view.columns:
                            view[c] = view[c].astype(str).fillna("").map(lambda x: (x[:80] + "…") if len(x) > 81 else x)
                    st.dataframe(view, use_container_width=True, hide_index=True, height=320)

    if status_filter:
        orders = orders_all[orders_all["status"].isin(status_filter)]
        items = items_all[items_all["order_id"].isin(orders["order_id"])]
    else:
        orders = orders_all.copy()
        items = items_all.copy()

    today = dt.date.today()
    if date_basis.startswith("purchase_date") and "purchase_date" in orders.columns:
        orders["_date"] = _to_date_series(orders["purchase_date"])
    else:
        orders["_date"] = _to_date_series(orders["created_at"])

    ship_tbl = claim_tbl = back_tbl = close_tbl = cancel_tbl = False

    # Build the shared "접수 목록" table frames (used by page1 + sidebar detail editor)
    user_cols = [
        "purchase_date",
        "receiver_name",
        "동일연락처",
        "배송(엑셀)",
        "order_list",
        "address",
        "phone",
        "delivery_request",
        "attention_note",
        "special_issue",
        "status",
    ]
    admin_cols = [
        "order_id",
        "party_key",
        "source_file",
        "group_no",
        "deadline_raw",
        "order_date_raw",
        "created_at",
    ]
    all_cols = user_cols + (admin_cols if show_admin_cols else [])
    # 전체 접수 목록은 "최근 날짜가 위"로 정렬 (원본 엑셀 파일명 순이 아니라 업무 흐름 기준)
    sorted_orders = orders.copy()
    sort_cols: list[str] = []
    ascending: list[bool] = []
    if "_date" in sorted_orders.columns:
        sort_cols.append("_date")
        ascending.append(False)
    if "created_at" in sorted_orders.columns:
        sort_cols.append("created_at")
        ascending.append(False)
    # 동률일 때는 파일/그룹 순으로 안정적으로
    for c in ["source_file", "group_no"]:
        if c in sorted_orders.columns:
            sort_cols.append(c)
            ascending.append(True)
    if sort_cols:
        sorted_orders = sorted_orders.sort_values(sort_cols, ascending=ascending, na_position="last")
    sorted_orders = sorted_orders.drop(columns=["_date"], errors="ignore")
    if "order_id" in sorted_orders.columns and len(items_all) and "order_id" in items_all.columns:
        _ship_map = infer_settlement_ship_series(items_all)
        sorted_orders = sorted_orders.copy()
        sorted_orders["배송(엑셀)"] = (
            sorted_orders["order_id"].astype(str).map(_ship_map).fillna("택배").map(_ship_display_label)
        )
    if "party_key" in sorted_orders.columns:
        _pk = sorted_orders["party_key"].astype(str)
        _grp_n = _pk.groupby(_pk).transform("count")
        sorted_orders["동일연락처"] = _grp_n.map(lambda x: f"{int(x)}건" if int(x) > 1 else "—")
    else:
        sorted_orders = sorted_orders.copy()
        sorted_orders["동일연락처"] = "—"
    view_orders = sorted_orders[[c for c in all_cols if c in sorted_orders.columns]].copy()

    # Make status visually distinct (emoji badges). Streamlit's dataframe has limited per-row styling.
    if "status" in view_orders.columns:
        def _badge(s: object) -> str:
            ss = "" if s is None else str(s).strip()
            if ss == "출고":
                return "🚚✅ 출고완료"
            if ss == "접수":
                return "📝⏳ 접수"
            if ss == "도면참조":
                return "📐 도면참조"
            if ss == "클레임":
                return "⚠️ 클레임"
            if ss == "마감":
                return "💰✅ 마감"
            if ss == "납품취소":
                return "⛔ 납품취소"
            return ss or ""

        view_orders["status"] = view_orders["status"].map(_badge)

    # 한 칸: 상태 이모지 1 + 이름(최대 8자) + 뒤 아이콘 최대 3 (status 열은 긴 뱃지 유지).
    if "receiver_name" in view_orders.columns and "status" in sorted_orders.columns:
        def _name_emoji(row) -> str:
            name = "" if row.get("receiver_name") is None else str(row.get("receiver_name")).strip()
            stt = "" if row.get("status") is None else str(row.get("status")).strip()
            att_src = _combined_attention_for_icons(row.get("attention_note"), row.get("special_issue"))
            blob = _drawing_ref_sheet_blob(row, items_all)
            return _compact_name_display(stt, name, att_src, sheet_blob=blob)

        # Use sorted_orders for original status values
        cols_tmp = [
            c
            for c in [
                "receiver_name",
                "status",
                "attention_note",
                "special_issue",
                "order_list",
                "order_id",
                "delivery_request",
                "deadline_raw",
            ]
            if c in sorted_orders.columns
        ]
        tmp = sorted_orders.loc[view_orders.index, cols_tmp].copy()
        view_orders["receiver_name"] = tmp.apply(_name_emoji, axis=1)

    # Page routing (sidebar buttons)
    page = str(st.session_state.get("mutomo_page") or "page1")
    selected_ids: list[str] = []

    if page == "page1":
        pend = st.session_state.get(MUTOMO_PENDING_REVERT_KEY)
        if pend and isinstance(pend, dict):
            ids = [str(x) for x in (pend.get("order_ids") or [])]
            wmsg = str(pend.get("message") or "").strip()
            src = str(pend.get("source") or "")
            src_ko = "이름 검색에서 고른 주문" if src == "search" else "접수 목록에서 고른 주문"
            st.warning(
                f"**접수로 되돌리기 확인** ({src_ko})\n\n"
                f"- 요약: {wmsg if wmsg else '이미 처리 이력이 있습니다.'}\n\n"
                "계속하면 **출고시각·마감시각**이 비워지고 상태가 **접수**로 바뀝니다."
            )
            c_can, c_ok = st.columns(2)
            with c_can:
                if st.button("취소", key="mutomo_pending_revert_cancel"):
                    st.session_state.pop(MUTOMO_PENDING_REVERT_KEY, None)
                    st.rerun()
            with c_ok:
                if st.button("그래도 접수로 변경", type="primary", key="mutomo_pending_revert_confirm"):
                    _dbp = str(st.session_state.get("mutomo_db_path", "mutomo.sqlite"))
                    if ids:
                        con = sqlite3.connect(_dbp)
                        try:
                            cur = con.cursor()
                            cur.executemany(
                                "UPDATE orders SET status=?, shipped_at=NULL, closed_at=NULL WHERE order_id=?",
                                [("접수", oid) for oid in ids],
                            )
                            con.commit()
                        finally:
                            con.close()
                    st.session_state.pop(MUTOMO_PENDING_REVERT_KEY, None)
                    st.success(f"접수로 변경했습니다: {len(ids)}건")
                    st.cache_data.clear()
                    st.rerun()
            st.divider()
        # page1: 접수목록
        st.subheader("접수목록")
        # 이름 검색은 사이드바 상태 필터와 무관하게 "전체 주문"에서 찾는 게 자연스럽다.
        render_receiver_name_search(orders_all, items_all, db_path, orders_name_hints=orders_all)
        st.divider()
        st.subheader("전체 접수 목록")
        if len(orders) > 5000:
            st.caption(
                f"현재 표에 올리는 행이 **{len(orders):,}건**입니다. "
                "느려지면 상태 필터로 줄이거나, ingest 전에 SQLite에서 기간·상태로 나눠 저장하는 방식을 검토하세요."
            )
        st.caption(
            "🟥 표시는 **특이사항(자동)** 또는 **특이사항(수기)** 이 있는 주문입니다. (이름 열 아이콘에 둘 다 반영됩니다.) "
            "**배송(엑셀)** 은 목록에서만 보는 품목 배송란 다수결 요약입니다. "
            "사이드바 상세의 배송은 **품목** 아래 **배송:** 에 적힌 엑셀 원문만 쓰며, 변경은 현장에서 처리합니다. "
            "**동일연락처** 는 전화(또는 주소)로 묶은 그룹 안의 접수 건수(재주문·동일 수하인 힌트)이며, "
            "확정 식별은 항상 **order_id**(한 줄 한 접수)입니다."
        )

        st.caption(
            "행을 고르면 왼쪽 사이드바에 상세가 뜹니다. "
            "**도면참조**는 엑셀 본문·이름 검색에서 반영됩니다."
        )
        b0, b1, b2, b3, b4, b5 = st.columns(6, gap="small")
        with b0:
            clear_pick = st.button("선택 해제", key="table_clear_selection_btn", help="표 선택(체크)을 세션에서 초기화합니다.")
        with b1:
            ship_tbl = st.button("선택 출고", type="primary", key="table_ship_btn")
        with b2:
            claim_tbl = st.button("클레임", key="table_claim_btn")
        with b3:
            back_tbl = st.button("접수", key="table_received_btn")
        with b4:
            close_tbl = st.button("마감", key="table_close_btn")
        with b5:
            cancel_tbl = st.button("납품취소", key="table_cancel_btn")

        name_col_cfg: dict | None = None
        _cols_cfg: dict = {}
        if "receiver_name" in view_orders.columns:
            _cols_cfg["receiver_name"] = st.column_config.TextColumn("이름", width="medium")
        if "동일연락처" in view_orders.columns:
            _cols_cfg["동일연락처"] = st.column_config.TextColumn("동일연락처", width="small")
        if "배송(엑셀)" in view_orders.columns:
            _cols_cfg["배송(엑셀)"] = st.column_config.TextColumn("배송(엑셀)", width="small")
        if "party_key" in view_orders.columns:
            _cols_cfg["party_key"] = st.column_config.TextColumn("party_key", width="small")
        if _cols_cfg:
            name_col_cfg = _cols_cfg

        if "order_id" in sorted_orders.columns:
            _df_sel_key = "mutomo_full_orders_df"
            if clear_pick:
                st.session_state["selected_ids_from_table"] = []
                st.session_state["prev_table_pick_ids"] = []
                st.session_state["active_selector"] = None
                st.session_state["table_sync_from_search"] = False
                if _df_sel_key in st.session_state:
                    st.session_state[_df_sel_key] = {"selection": {"rows": [], "columns": [], "cells": []}}
                st.rerun()

            if st.session_state.pop("_search_sync_need_df_apply", False):
                sync_ids = st.session_state.get("_search_sync_order_ids") or []
                idset = set(sync_ids)
                row_idxs = [i for i in range(len(sorted_orders)) if str(sorted_orders.iloc[i]["order_id"]) in idset]
                st.session_state[_df_sel_key] = {"selection": {"rows": row_idxs, "columns": [], "cells": []}}
            else:
                # Streamlit dataframe selection can briefly reset across reruns, or row indices can drift if the
                # table order changes. Re-align widget row selection with persisted order_ids in `view_orders`.
                if (
                    (not st.session_state.get("table_sync_from_search"))
                    and st.session_state.get("active_selector") == "table"
                    and "order_id" in view_orders.columns
                ):
                    wanted = [str(x) for x in (st.session_state.get("selected_ids_from_table") or []) if str(x)]
                    if wanted:
                        want = set(wanted)
                        row_idxs2: list[int] = []
                        for i in range(len(view_orders)):
                            oid = str(view_orders.iloc[i].get("order_id") or "")
                            if oid in want:
                                row_idxs2.append(i)
                        prev_rows: list[int] = []
                        _prev = st.session_state.get(_df_sel_key)
                        if isinstance(_prev, dict):
                            sel = _prev.get("selection") or {}
                            pr = sel.get("rows")
                            if isinstance(pr, list):
                                for x in pr:
                                    try:
                                        xi = int(x)
                                    except (TypeError, ValueError):
                                        continue
                                    prev_rows.append(xi)
                        if prev_rows != row_idxs2:
                            st.session_state[_df_sel_key] = {"selection": {"rows": row_idxs2, "columns": [], "cells": []}}
            try:
                state = st.dataframe(
                    view_orders,
                    use_container_width=True,
                    hide_index=True,
                    column_config=name_col_cfg,
                    on_select="rerun",
                    selection_mode="multi-row",
                    key=_df_sel_key,
                )
                if state is not None and hasattr(state, "selection"):
                    rows = getattr(state.selection, "rows", []) or []
                    # selection.rows는 표시된 표 기준 위치인데, 필터·재수집 후 행 수가 줄면 이전 인덱스가 남아 IndexError가 난다.
                    n_vis = len(view_orders)
                    valid_rows: list[int] = []
                    for x in rows:
                        try:
                            xi = int(x)
                        except (TypeError, ValueError):
                            continue
                        if 0 <= xi < n_vis:
                            valid_rows.append(xi)
                    if valid_rows:
                        # view_orders는 화면 표시용이라 order_id 컬럼이 없을 수 있음(관리 컬럼 숨김).
                        # 대신 view_orders의 인덱스를 기준으로 sorted_orders에서 order_id를 뽑는다.
                        base = sorted_orders.loc[view_orders.index]
                        if "order_id" in base.columns:
                            selected_ids = base.iloc[valid_rows]["order_id"].astype(str).tolist()
                    elif rows:
                        # 표가 줄거나 정렬이 바뀌는 순간 Streamlit이 잠깐 잘못된 row index를 보낼 수 있다.
                        # 이 때 선택을 "꺼버리면" UX가 나빠서, 기존에 저장된 order_id 선택을 유지한다.
                        selected_ids = [str(x) for x in (st.session_state.get("selected_ids_from_table") or []) if str(x)]
            except TypeError:
                # Older Streamlit: no selection support, just show the table.
                st.dataframe(view_orders, use_container_width=True, hide_index=True, column_config=name_col_cfg)
        else:
            st.dataframe(view_orders, use_container_width=True, hide_index=True, column_config=name_col_cfg)

    elif page == "page2":
        # page2: 출고
        st.subheader("출고")
        with st.container():

            def _sales_section_bar() -> None:
                st.markdown(
                    '<div style="height:3px;background:linear-gradient(90deg,transparent,#90A4AE,#37474F,#90A4AE,transparent);'
                    'border-radius:2px;margin:1.35rem 0 0.9rem 0;opacity:0.95;"></div>',
                    unsafe_allow_html=True,
                )

            # Right-side summary metrics
            s = orders_all["status"] if "status" in orders_all.columns else pd.Series([], dtype=str)
            total_cnt = len(orders_all)
            planned_cnt = int((s == "접수").sum()) if len(s) else 0
            custom_cnt = int((s == "도면참조").sum()) if len(s) else 0
            done_cnt = int((s == "출고").sum()) if len(s) else 0
            claim_cnt = int((s == "클레임").sum()) if len(s) else 0
            closed_cnt = int((s == "마감").sum()) if len(s) else 0
            cancel_cnt = int((s == "납품취소").sum()) if len(s) else 0

            left, m1, m2, m3, m4, m5, m6, m7 = st.columns([1.7, 1, 1, 1, 1, 1, 1, 1])
            with left:
                st.metric("오늘 접수(주문그룹)", int((orders["_date"] == today).sum()))
            with m1:
                st.metric("전체접수", total_cnt)
            with m2:
                st.metric("납품예정", planned_cnt)
            with m3:
                st.metric("도면참조", custom_cnt)
            with m4:
                st.metric("납품완료", done_cnt)
            with m5:
                st.metric("마감", closed_cnt)
            with m6:
                st.metric("클레임", claim_cnt)
            with m7:
                st.metric("납품취소", cancel_cnt)

            _sales_section_bar()
            st.subheader("오늘접수")
            st.caption("**접수**와 **도면참조**(도면 주고 제작) 모두 이 패널에 포함됩니다.")

            # 접수·도면참조만 일별 패널에 표시 (그 외는 접수목록에서 확인)
            recent_base = orders.copy()
            if "status" in recent_base.columns:
                recent_base = recent_base[recent_base["status"].isin(["접수", "도면참조"])]

            def _day_panel(col, day: dt.date) -> None:
                df = recent_base[recent_base["_date"] == day].copy()
                n_actual = len(df)
                # Show request/attention icons in the "today" panels too
                keep = [
                    c
                    for c in [
                        "purchase_date",
                        "receiver_name",
                        "attention_note",
                        "special_issue",
                        "order_list",
                        "status",
                        "order_id",
                        "delivery_request",
                        "deadline_raw",
                    ]
                    if c in df.columns
                ]
                df = df[keep]
                if "purchase_date" not in df.columns:
                    df["purchase_date"] = day.isoformat()
                if "receiver_name" not in df.columns:
                    df["receiver_name"] = ""
                if "attention_note" not in df.columns:
                    df["attention_note"] = ""
                if "special_issue" not in df.columns:
                    df["special_issue"] = ""
                if "order_list" not in df.columns:
                    df["order_list"] = ""
                if "status" not in df.columns:
                    df["status"] = ""
                if "order_id" not in df.columns:
                    df["order_id"] = ""
                if "delivery_request" not in df.columns:
                    df["delivery_request"] = ""
                if "deadline_raw" not in df.columns:
                    df["deadline_raw"] = ""

                def _disp_name(r: pd.Series) -> str:
                    nm = "" if r.get("receiver_name") is None else str(r.get("receiver_name")).strip()
                    stt = "" if r.get("status") is None else str(r.get("status")).strip()
                    att_src = _combined_attention_for_icons(r.get("attention_note"), r.get("special_issue"))
                    return _compact_name_display(
                        stt, nm, att_src, sheet_blob=_drawing_ref_sheet_blob(r, items_all)
                    )

                df["_이름표시"] = df.apply(_disp_name, axis=1)
                df = df.rename(columns={"purchase_date": "날짜"})
                df = df[["날짜", "_이름표시"]].rename(columns={"_이름표시": "이름"}).sort_values(["이름"], na_position="last")
                h_date, h_qty = col.columns([5, 2])
                with h_date:
                    st.markdown(f"**{day.isoformat()}**")
                with h_qty:
                    st.markdown(f"**수량 {n_actual}**")
                col.dataframe(
                    df,
                    use_container_width=True,
                    hide_index=True,
                    column_config={"이름": st.column_config.TextColumn("이름", width="medium")},
                )

            c1, c2, c3, c4 = st.columns(4)
            _day_panel(c1, today)
            _day_panel(c2, today - dt.timedelta(days=1))
            _day_panel(c3, today - dt.timedelta(days=2))
            _day_panel(c4, today - dt.timedelta(days=3))

            _sales_section_bar()
            st.subheader("최근 출고")
            # "최근 출고"는 사이드바 상태 필터(접수만 보기 등)와 무관하게 항상 출고 건을 보여준다.
            ship_base = orders_all.copy()
            if "status" in ship_base.columns and len(ship_base):
                ship_base = ship_base[ship_base["status"].astype(str).str.strip() == "출고"].copy()
            else:
                ship_base = orders_all.iloc[:0].copy()
            if len(ship_base) and "shipped_at" in ship_base.columns:
                ship_base["_ship_day"] = _shipped_at_calendar_date_series(ship_base["shipped_at"])
            else:
                ship_base["_ship_day"] = pd.Series(dtype=object)
            n_ship_no_time = 0
            if len(ship_base) and "_ship_day" in ship_base.columns:
                n_ship_no_time = int(ship_base["_ship_day"].isna().sum())
            if n_ship_no_time:
                st.caption(f"⚠️ 출고 {n_ship_no_time}건은 `shipped_at`이 비어 있어 날짜별 패널에서 제외될 수 있습니다. (출고 버튼으로 다시 저장하면 자동 보정됩니다.)")

            def _ship_day_panel(col, day: dt.date, idx: int) -> None:
                if len(ship_base) and "_ship_day" in ship_base.columns:
                    df = ship_base[ship_base["_ship_day"] == day].copy()
                else:
                    df = ship_base.iloc[:0].copy()
                n_ship = len(df)
                keep = [
                    c
                    for c in [
                        "shipped_at",
                        "receiver_name",
                        "attention_note",
                        "special_issue",
                        "order_list",
                        "status",
                        "order_id",
                        "delivery_request",
                        "deadline_raw",
                        "address",
                    ]
                    if c in df.columns
                ]
                df = df[keep] if len(keep) else pd.DataFrame()
                if "shipped_at" not in df.columns:
                    df["shipped_at"] = pd.NaT
                if "receiver_name" not in df.columns:
                    df["receiver_name"] = ""
                if "address" not in df.columns:
                    df["address"] = ""
                if "attention_note" not in df.columns:
                    df["attention_note"] = ""
                if "special_issue" not in df.columns:
                    df["special_issue"] = ""
                if "order_list" not in df.columns:
                    df["order_list"] = ""
                if "status" not in df.columns:
                    df["status"] = ""
                if "order_id" not in df.columns:
                    df["order_id"] = ""
                if "delivery_request" not in df.columns:
                    df["delivery_request"] = ""
                if "deadline_raw" not in df.columns:
                    df["deadline_raw"] = ""

                def _disp_ship_name(r: pd.Series) -> str:
                    nm = "" if r.get("receiver_name") is None else str(r.get("receiver_name")).strip()
                    stt = "" if r.get("status") is None else str(r.get("status")).strip()
                    att_c = _strip_order_list_overlap(r.get("attention_note"), r.get("order_list"))
                    att_src = _combined_attention_for_icons(att_c or None, r.get("special_issue"))
                    return _compact_name_display(
                        stt, nm, att_src, sheet_blob=_drawing_ref_sheet_blob(r, items_all)
                    )

                df["_이름표시"] = df.apply(_disp_ship_name, axis=1)
                ts = pd.to_datetime(df["shipped_at"], errors="coerce")
                df["_출고일시"] = ts.dt.strftime("%Y-%m-%d %H:%M").fillna("")
                df["_주소"] = df["address"].astype(str).fillna("").map(lambda x: x.strip().replace("\n", " "))
                df["_주소"] = df["_주소"].map(lambda x: (x[:32] + "…") if len(x) > 33 else x)
                df = (
                    df[["_출고일시", "_이름표시", "_주소"]]
                    .rename(columns={"_출고일시": "출고", "_이름표시": "이름", "_주소": "주소"})
                    .sort_values(["이름"], na_position="last")
                )
                h_sd, h_sq = col.columns([5, 2])
                with h_sd:
                    st.markdown(f"**{day.isoformat()}**")
                with h_sq:
                    st.markdown(f"**수량 {n_ship}**")
                col.dataframe(
                    df,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "출고": st.column_config.TextColumn("출고", width="small"),
                        "이름": st.column_config.TextColumn("이름", width="medium"),
                        "주소": st.column_config.TextColumn("주소", width="large"),
                    },
                )

            # Always show today → 3 days ago (requested UX).
            days: list[dt.date] = [
                today,
                today - dt.timedelta(days=1),
                today - dt.timedelta(days=2),
                today - dt.timedelta(days=3),
            ]

            cols = st.columns(4)
            for j in range(4):
                _ship_day_panel(cols[j], days[j], j)

            # shipped_at이 비어 날짜로 묶이지 않는 출고건도 별도로 보여준다
            if n_ship_no_time and len(ship_base):
                miss = ship_base[ship_base["_ship_day"].isna()].copy()
                keep2 = [c for c in ["receiver_name", "order_list", "order_id"] if c in miss.columns]
                miss = miss[keep2] if keep2 else pd.DataFrame()
                if "receiver_name" in miss.columns:
                    miss["받는분"] = miss["receiver_name"].astype(str).str.strip()
                if "order_list" in miss.columns:
                    miss["주문"] = miss["order_list"].astype(str).fillna("").map(lambda x: (x[:80] + "…") if len(x) > 81 else x)
                cols_m = [c for c in ["받는분", "주문", "order_id"] if c in miss.columns]
                if cols_m:
                    st.subheader("출고시각 없음(보정 필요)")
                    st.dataframe(miss[cols_m].head(60), use_container_width=True, hide_index=True, height=240)

    else:
        # page3: 마감
        render_sales_period_tab(orders_all, items_all)

    # Mirror "name search" behavior: show selected rows in sidebar
    # Persist selection so button clicks don't lose it on rerun
    if selected_ids:
        if selected_ids != st.session_state.get("prev_table_pick_ids"):
            st.session_state["prev_table_pick_ids"] = selected_ids
            st.session_state["active_selector"] = "table"
            st.session_state["selected_ids_from_table"] = selected_ids
            st.session_state["table_sync_from_search"] = False
            # Clear search selection (including widget state) if table was used
            st.session_state["search_pick_ids"] = []
            st.session_state["prev_search_pick_ids"] = []
            st.session_state["request_clear_search_pick_labels"] = True

    selected_ids_persisted = st.session_state.get("selected_ids_from_table", [])

    if selected_ids_persisted and st.session_state.get("active_selector") == "table":
        # order_id dtype가 섞여도(숫자/문자) 선택 매칭이 되도록 문자열로 통일
        idset = {str(x) for x in (selected_ids_persisted or []) if str(x)}
        # 사이드바 상세는 화면 필터에 영향받지 않게 "전체 주문"에서 가져온다.
        if "order_id" in orders_all.columns and idset:
            picked = orders_all[orders_all["order_id"].astype(str).isin(list(idset))].copy()
        else:
            picked = orders_all.iloc[:0].copy()
        sort_cols_pick = [c for c in ["purchase_date", "receiver_name"] if c in picked.columns]
        if sort_cols_pick:
            picked = picked.sort_values(sort_cols_pick, na_position="last")
        st.sidebar.subheader("접수목록 선택 상세")
        st.sidebar.caption(f"선택: {len(selected_ids_persisted)}건")
        for _, r in picked.iterrows():
            _render_order_detail(st.sidebar, r, items_all)
            st.sidebar.divider()

    # Claim details / special notes editor (applies to current selection)
    active_ids: list[str] = []
    active_label = ""
    if st.session_state.get("active_selector") == "search":
        active_ids = st.session_state.get("search_pick_ids", []) or []
        active_label = "검색 선택"
    elif st.session_state.get("active_selector") == "table":
        active_ids = selected_ids_persisted or []
        active_label = "접수목록 선택"

    if active_ids:
        st.sidebar.subheader("클레임/특이사항")
        if len(active_ids) > 1:
            st.sidebar.caption(f"{active_label}: {len(active_ids)}건 (한 번에 동일 내용 저장)")
        else:
            st.sidebar.caption(f"{active_label}: 1건")

        # Pre-fill from the first selected order
        first_row = orders_all[orders_all["order_id"].astype(str) == str(active_ids[0])]
        current_text = ""
        if len(first_row) and "special_issue" in first_row.columns:
            v = first_row.iloc[0].get("special_issue")
            current_text = "" if v is None else str(v)

        issue_text = st.sidebar.text_area(
            "내용",
            value=current_text,
            height=110,
            key=f"special_issue_editor_{st.session_state.get('active_selector')}_{str(active_ids[0])}",
            placeholder="예) 오염/파손, 색상 변경 요청, 부분 환불, 재발송 필요 등",
        )
        if st.sidebar.button("특이사항 저장", type="secondary", key="save_special_issue"):
            update_special_issue_for_orders(db_path, active_ids, issue_text.strip())
            st.sidebar.success("저장했습니다.")
            st.cache_data.clear()
            st.rerun()

    else:
        st.sidebar.caption(
            "**이름 검색** 탭 오른쪽 **출고할 주문 선택**에서 고르거나, 아래 **전체 접수 목록**에서 행을 고르면 "
            "**클레임/특이사항** 메모를 저장할 수 있습니다. "
            "목록의 **배송(엑셀)** 은 스캔용 요약이며, 상세 배송은 품목 줄의 엑셀 원문만 봅니다."
        )

    # 목록 선택 전용: 이름 검색 탭의 출고 버튼과 분리된 키를 씁니다.
    if ship_tbl and st.session_state.get("active_selector") == "table":
        if selected_ids_persisted:
            con = sqlite3.connect(db_path)
            try:
                cur = con.cursor()
                now = dt.datetime.now().isoformat(timespec="seconds")
                cur.executemany(
                    "UPDATE orders SET status=?, shipped_at=COALESCE(NULLIF(shipped_at,''), ?), closed_at=NULL WHERE order_id=?",
                    [("출고", now, oid) for oid in selected_ids_persisted],
                )
                con.commit()
            finally:
                con.close()
            st.sidebar.success(f"출고 처리: {len(selected_ids_persisted)}건")
            st.cache_data.clear()
            st.rerun()

    if claim_tbl and st.session_state.get("active_selector") == "table":
        if selected_ids_persisted:
            con = sqlite3.connect(db_path)
            try:
                cur = con.cursor()
                cur.executemany(
                    "UPDATE orders SET status=?, closed_at=NULL WHERE order_id=?",
                    [("클레임", oid) for oid in selected_ids_persisted],
                )
                con.commit()
            finally:
                con.close()
            st.sidebar.success(f"클레임 처리: {len(selected_ids_persisted)}건")
            st.cache_data.clear()
            st.rerun()

    if back_tbl and st.session_state.get("active_selector") == "table":
        if selected_ids_persisted:
            if "order_id" in orders_all.columns:
                sub0 = orders_all[orders_all["order_id"].astype(str).isin([str(x) for x in selected_ids_persisted])].copy()
            else:
                sub0 = pd.DataFrame()
            need_warn, wmsg = _revert_to_received_warn_message(sub0)
            if need_warn:
                st.session_state[MUTOMO_PENDING_REVERT_KEY] = {
                    "order_ids": [str(x) for x in selected_ids_persisted],
                    "source": "table",
                    "message": wmsg,
                }
                st.rerun()
            con = sqlite3.connect(db_path)
            try:
                cur = con.cursor()
                cur.executemany(
                    "UPDATE orders SET status=?, shipped_at=NULL, closed_at=NULL WHERE order_id=?",
                    [("접수", oid) for oid in selected_ids_persisted],
                )
                con.commit()
            finally:
                con.close()
            st.sidebar.success(f"접수로 변경했습니다. (출고/마감 초기화): {len(selected_ids_persisted)}건")
            st.cache_data.clear()
            st.rerun()

    if close_tbl and st.session_state.get("active_selector") == "table":
        if selected_ids_persisted:
            con = sqlite3.connect(db_path)
            try:
                cur = con.cursor()
                cur.executemany(
                    "UPDATE orders SET status=?, closed_at=COALESCE(closed_at, ?) WHERE order_id=?",
                    [("마감", dt.datetime.now().isoformat(timespec="seconds"), oid) for oid in selected_ids_persisted],
                )
                con.commit()
            finally:
                con.close()
            st.sidebar.success(f"마감 처리: {len(selected_ids_persisted)}건")
            st.cache_data.clear()
            st.rerun()

    if cancel_tbl and st.session_state.get("active_selector") == "table":
        if selected_ids_persisted:
            con = sqlite3.connect(db_path)
            try:
                cur = con.cursor()
                cur.executemany(
                    "UPDATE orders SET status=?, closed_at=NULL WHERE order_id=?",
                    [("납품취소", oid) for oid in selected_ids_persisted],
                )
                con.commit()
            finally:
                con.close()
            st.sidebar.success(
                f"납품취소 처리: {len(selected_ids_persisted)}건. "
                "목록에서 안 보이면 **설정 → 상태 필터**에 **납품취소**가 포함돼 있는지 확인하세요."
            )
            st.cache_data.clear()
            st.rerun()

    with st.expander("제품 목록(집계)", expanded=False):
        st.caption(
            "위 **상태 필터**에 맞는 품목 라인만 집계합니다. "
            "정규명은 `product_aliases.yml` 매핑 뒤의 표준 상품명입니다. "
            "줄 단위 원본은 아래 **전체 품목(라인아이템) 보기**를 열어 확인하세요."
        )
        g_map, g_raw = _product_catalog_summary(items)
        if len(g_map) == 0 and len(g_raw) == 0:
            st.info("표시할 품목 라인이 없습니다.")
        else:
            if len(g_map):
                st.markdown("**정규 상품명** (매핑됨)")
                st.dataframe(g_map, use_container_width=True, hide_index=True)
            if len(g_raw):
                st.markdown("**엑셀 원문** (정규명 미매핑)")
                st.dataframe(g_raw, use_container_width=True, hide_index=True)

    with st.expander("전체 품목(라인아이템) 보기", expanded=False):
        show = items.copy()
        show["is_unmapped"] = show["product_canonical"].isna()
        st.dataframe(
            show.sort_values(["source_file", "order_id", "row_idx"], na_position="last"),
            use_container_width=True,
            hide_index=True,
        )

    with st.expander("상품명 매핑 작업(미매핑만)", expanded=False):
        unmapped = show[show["product_canonical"].isna() & show["product_key"].notna()].copy()
        if len(unmapped) == 0:
            st.info("현재 미매핑 상품이 없습니다.")
        else:
            st.warning("`product_aliases.yml`에 별칭을 추가하면 다음 수집부터 자동 매핑됩니다.")
            st.dataframe(
                unmapped[
                    [
                        "product_raw",
                        "product_key",
                        "suggested_canonical",
                        "suggestion_score",
                        "source_file",
                        "row_idx",
                    ]
                ]
                .drop_duplicates()
                .sort_values(["suggestion_score"], ascending=False, na_position="last"),
                use_container_width=True,
                hide_index=True,
            )


if __name__ == "__main__":
    main()

